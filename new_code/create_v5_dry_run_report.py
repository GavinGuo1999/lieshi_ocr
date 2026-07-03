import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


BASE = Path(r"D:\ying_min_mineru")
V4_XLSX = BASE / "英名录25版-祁县-二审_v4.xlsx"
FULL237_MANIFEST = BASE / "data" / "20260626" / "extracted_mineru_text_full237" / "mineru_text_manifest.json"
NEW6_MANIFEST = BASE / "data" / "20260626" / "extracted_mineru_text_new6" / "mineru_text_manifest.json"
OUT_MD = BASE / "log" / "20260630_v5_dry_run_report.md"
OUT_JSON = BASE / "log" / "20260630_v5_dry_run_report.json"


COLS = {
    "code": 2,
    "name": 4,
    "origin": 6,
    "birth": 7,
    "join": 8,
    "political": 9,
    "unit_role": 10,
    "story": 11,
    "story_backup": 20,
}

FIELD_TO_TARGET = {
    "姓名": "name",
    "籍贯": "origin",
    "出生时间": "birth",
    "出生日期": "birth",
    "参加革命时间": "join",
    "参加工作时间": "join",
    "参加革命（工作）时间": "join",
    "参加革命(工作)时间": "join",
    "政治面貌": "political",
    "生前单位及曾任职务": "unit_role",
    "生前（部队）单位及曾任职务": "unit_role",
    "生前(部队)单位及曾任职务": "unit_role",
    "曾任职务": "unit_role",
    "事迹": "story",
}

STORY_SOURCE_FIELDS = ("事迹", "牺牲时间", "牺牲地点", "牺牲原因")
UNMAPPED_FIELDS = {"民族", "安葬地", "安葬地点", "信息管理单位"}

ROLE_SUFFIXES = [
    "地下工作人员",
    "一般工作人员",
    "工作人员",
    "民兵指导员",
    "武委会副主任",
    "财粮主任",
    "财政主任",
    "副村长",
    "副主任",
    "副科长",
    "副部长",
    "副连长",
    "副排长",
    "副班长",
    "指导员",
    "通讯员",
    "通信员",
    "交通员",
    "情报员",
    "宣传员",
    "管理员",
    "采购员",
    "工作员",
    "看护员",
    "卫生员",
    "侦察员",
    "炊事员",
    "中医师",
    "司号员",
    "小队长",
    "大队长",
    "支队长",
    "书记",
    "主任",
    "科长",
    "区长",
    "部长",
    "连长",
    "排长",
    "班长",
    "队长",
    "村长",
    "处长",
    "干部",
    "战士",
    "队员",
    "民工",
    "民兵",
    "团员",
    "参谋",
    "文书",
    "医师",
    "村民",
]

TOWNS = ["昭馀镇", "古县镇", "峪口乡", "来远镇", "东观镇", "贾令镇", "城赵镇", "西六支乡", "麓台城区", "城区"]


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def comparable(value):
    text = cell_text(value)
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("，", ",").replace("；", ";").replace("。", "")
    return re.sub(r"\s+", "", text)


def normalize_date_value(value):
    text = cell_text(value)
    if not text:
        return text
    text = text.replace("０", "0").replace("１", "1").replace("２", "2").replace("３", "3").replace("４", "4")
    text = text.replace("５", "5").replace("６", "6").replace("７", "7").replace("８", "8").replace("９", "9")
    text = text.replace("年以前", "-前").replace("年前", "-前")

    m = re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日?", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2))}-{int(m.group(3))}"
    m = re.fullmatch(r"(\d{4})年(\d{1,2})月", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2))}"
    m = re.fullmatch(r"(\d{4})年", text)
    if m:
        return m.group(1)
    m = re.fullmatch(r"(\d{4})年-?前", text)
    if m:
        return f"{m.group(1)}-前"

    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def format_origin(value):
    text = cell_text(value)
    if not text:
        return text
    prefix = "山西省晋中市祁县"
    if not text.startswith(prefix):
        return text
    rest = text[len(prefix):].strip()
    if not rest:
        return prefix
    for town in sorted(TOWNS, key=len, reverse=True):
        if rest.startswith(town):
            village = rest[len(town):].strip("- ")
            return f"{prefix}\n{town}-{village}" if village else f"{prefix}\n{town}"
    return f"{prefix}\n{rest}"


def split_unit_role(value):
    text = cell_text(value)
    if not text or "/" in text:
        return text
    m = re.fullmatch(r"(.+?)[(（]([^()（）]+)[)）]", text)
    if m:
        return f"{m.group(1).strip()}/{m.group(2).strip()}"
    for role in sorted(ROLE_SUFFIXES, key=len, reverse=True):
        if text.endswith(role) and len(text) > len(role):
            unit = text[: -len(role)].strip()
            if unit:
                return f"{unit}/{role}"
    return text


def normalize_story_text(value):
    text = cell_text(value)
    if not text:
        return text
    text = re.sub(r"\s+", "", text)
    text = text.replace(",", "，").replace(";", "；")
    text = text.replace("壮烈牺牲", "牺牲").replace("光荣牺牲", "牺牲").replace("英勇牺牲", "牺牲")
    text = text.replace("中弹牺牲", "牺牲").replace("牺牲牺牲", "牺牲")
    return text


def ensure_period(text):
    text = cell_text(text).strip("，,；;。 ")
    return f"{text}。" if text else text


def simplify_story_from_items(items):
    sacrifice_time = cell_text(items.get("牺牲时间", {}).get("value"))
    sacrifice_place = cell_text(items.get("牺牲地点", {}).get("value"))
    sacrifice_reason = cell_text(items.get("牺牲原因", {}).get("value"))
    story = normalize_story_text(items.get("事迹", {}).get("value"))

    if sacrifice_reason:
        reason = normalize_story_text(sacrifice_reason)
        if sacrifice_time and not reason.startswith(sacrifice_time):
            reason = f"{sacrifice_time}{reason}"
        return ensure_period(reason)

    if sacrifice_time and sacrifice_place:
        return ensure_period(f"{sacrifice_time}在{sacrifice_place}牺牲")

    if story:
        event_words = [
            "牺牲", "病故", "病逝", "遇难", "阵亡", "被敌杀害", "被害", "枪杀", "炸死", "淹死",
            "与家庭失去联系", "失去联系", "下落不明", "无消息", "无音信",
        ]
        hit_positions = [story.find(word) for word in event_words if story.find(word) >= 0]
        if hit_positions:
            hit = min(hit_positions)
            date_matches = list(re.finditer(r"\d{4}年(?:\d{1,2}月)?(?:\d{1,2}日)?", story[: hit + 20]))
            start = date_matches[-1].start() if date_matches else max(story.rfind("。", 0, hit), story.rfind("；", 0, hit), story.rfind("，", 0, hit)) + 1
            end_candidates = [i for i in [story.find("。", hit), story.find("；", hit)] if i >= 0]
            end = min(end_candidates) if end_candidates else len(story)
            phrase = story[start:end]
            phrase = re.split(r"时年\d+岁", phrase)[0]
            if len(phrase) > 70:
                phrase = phrase[:70].rstrip("，,；;")
            return ensure_period(phrase)
    return ""


def normalize_for_target(target, value, items):
    if target in {"birth", "join"}:
        return normalize_date_value(value)
    if target == "origin":
        return format_origin(value)
    if target == "unit_role":
        return split_unit_role(value)
    if target == "story":
        return simplify_story_from_items(items)
    return cell_text(value)


def read_record_json(record):
    return load_json(Path(record["json"]))


def load_records():
    records = []
    for source_name, manifest_path in [("full237", FULL237_MANIFEST), ("new6", NEW6_MANIFEST)]:
        manifest = load_json(manifest_path)
        for rec in manifest["records"]:
            item = dict(rec)
            item["source_manifest_group"] = source_name
            records.append(item)
    return records


def build_excel_index(ws):
    by_code = {}
    duplicates = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        code = cell_text(ws.cell(row, COLS["code"]).value)
        if not code:
            continue
        if code in by_code:
            duplicates[code].append(row)
        else:
            by_code[code] = row
    return by_code, duplicates


def add_change(changes, row, code, name, target, field, old_value, new_value, reason, source_stem):
    if not cell_text(new_value):
        return
    if comparable(old_value) == comparable(new_value):
        return
    changes.append(
        {
            "row": row,
            "code": code,
            "name": name,
            "column": COLS[target],
            "target": target,
            "field": field,
            "old": cell_text(old_value),
            "new": cell_text(new_value),
            "reason": cell_text(reason),
            "source_stem": source_stem,
        }
    )


def main():
    records = load_records()
    wb = openpyxl.load_workbook(V4_XLSX, data_only=False, read_only=True)
    ws = wb.active
    code_to_row, excel_duplicates = build_excel_index(ws)

    code_counts = Counter(r["code"] for r in records)
    duplicate_record_codes = {code: count for code, count in code_counts.items() if count > 1}
    quality_counts = Counter(r.get("quality", "") for r in records)
    warning_counts = Counter(w for r in records for w in r.get("warnings", []))

    changes = []
    review_needed = []
    rows_not_found = []
    name_mismatches = []
    unmapped_items = []
    skipped_low_confidence_changes = []
    empty_items = []
    backup_populated = 0
    item_count_total = 0

    for record in records:
        code = record["code"]
        name = record["name"]
        row = code_to_row.get(code)
        detail = read_record_json(record)
        items = detail.get("correction_items", {})
        item_count_total += len(items)

        if record.get("quality") != "ok" or record.get("warnings"):
            review_needed.append(
                {
                    "code": code,
                    "name": name,
                    "source_stem": record.get("source_stem"),
                    "quality": record.get("quality"),
                    "warnings": record.get("warnings", []),
                    "items": len(items),
                    "markers": record.get("mineru_marker_count"),
                }
            )
        if not items:
            empty_items.append({"code": code, "name": name, "source_stem": record.get("source_stem")})

        if row is None:
            rows_not_found.append({"code": code, "name": name, "source_stem": record.get("source_stem")})
            continue

        excel_name = cell_text(ws.cell(row, COLS["name"]).value)
        if comparable(excel_name) != comparable(name):
            name_mismatches.append({"row": row, "code": code, "excel_name": excel_name, "ocr_name": name})

        if cell_text(ws.cell(row, COLS["story_backup"]).value):
            backup_populated += 1

        story_fields_present = any(field in items for field in STORY_SOURCE_FIELDS)
        for field, item in items.items():
            target = FIELD_TO_TARGET.get(field)
            if not target:
                if field not in STORY_SOURCE_FIELDS:
                    unmapped_items.append(
                        {
                            "row": row,
                            "code": code,
                            "name": name,
                            "field": field,
                            "value": cell_text(item.get("value")),
                            "reason": cell_text(item.get("reason")),
                        }
                    )
                continue
            if target == "story":
                continue
            new_value = normalize_for_target(target, item.get("value"), items)
            if target == "origin" and not origin_value_looks_complete(new_value):
                skipped_low_confidence_changes.append(
                    {
                        "row": row,
                        "code": code,
                        "name": name,
                        "field": field,
                        "old": cell_text(ws.cell(row, COLS[target]).value),
                        "new": cell_text(new_value),
                        "reason": "籍贯缺少明确乡镇/城区信息，dry-run 不自动列为拟改",
                    }
                )
                continue
            old_value = ws.cell(row, COLS[target]).value
            if target == "unit_role" and not unit_role_change_is_worth_reporting(old_value, new_value):
                skipped_low_confidence_changes.append(
                    {
                        "row": row,
                        "code": code,
                        "name": name,
                        "field": field,
                        "old": cell_text(old_value),
                        "new": cell_text(new_value),
                        "reason": "J列已具备可用 / 结构或候选未形成 / 结构，dry-run 不自动列为拟改",
                    }
                )
                continue
            add_change(changes, row, code, name, target, field, old_value, new_value, item.get("reason"), record.get("source_stem"))

        if story_fields_present:
            new_story = normalize_for_target("story", "", items)
            old_story = ws.cell(row, COLS["story"]).value
            if new_story and story_needs_simplification(old_story):
                add_change(
                    changes,
                    row,
                    code,
                    name,
                    "story",
                    "事迹/牺牲时间地点原因",
                    old_story,
                    new_story,
                    "由 MinerU 结构化项 dry-run 生成简要事迹",
                    record.get("source_stem"),
                )

    changes_by_col = Counter(str(c["column"]) for c in changes)
    changes_by_field = Counter(c["field"] for c in changes)

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input": {
            "v4_xlsx": str(V4_XLSX),
            "full237_manifest": str(FULL237_MANIFEST),
            "new6_manifest": str(NEW6_MANIFEST),
        },
        "summary": {
            "records_total": len(records),
            "correction_item_total": item_count_total,
            "expected_records_total": 243,
            "excel_rows": ws.max_row,
            "quality_counts": dict(quality_counts),
            "warning_counts": dict(warning_counts),
            "record_duplicate_codes": duplicate_record_codes,
            "excel_duplicate_codes": {k: v for k, v in excel_duplicates.items()},
            "rows_not_found": len(rows_not_found),
            "name_mismatches": len(name_mismatches),
            "review_needed": len(review_needed),
            "empty_items": len(empty_items),
            "backup_populated_for_matched_records": backup_populated,
            "proposed_changes": len(changes),
            "proposed_changes_by_column": dict(changes_by_col),
            "proposed_changes_by_field": dict(changes_by_field),
            "unmapped_items": len(unmapped_items),
            "skipped_low_confidence_changes": len(skipped_low_confidence_changes),
        },
        "rows_not_found": rows_not_found,
        "name_mismatches": name_mismatches,
        "review_needed": review_needed,
        "unmapped_items": unmapped_items,
        "skipped_low_confidence_changes": skipped_low_confidence_changes,
        "proposed_changes": changes,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_MD.write_text(render_markdown(report), encoding="utf-8")
    print(f"json: {OUT_JSON}")
    print(f"markdown: {OUT_MD}")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))


def md_escape(value):
    text = cell_text(value).replace("\n", "<br>")
    return text.replace("|", "\\|")


def origin_value_looks_complete(value):
    text = cell_text(value)
    if not text:
        return False
    if not text.startswith("山西省晋中市祁县") and not text.startswith("山西省祁县"):
        return True
    return any(token in text for token in ["镇", "乡", "城区"])


def story_needs_simplification(value):
    text = normalize_story_text(value)
    if not text:
        return True
    if len(text) > 85:
        return True
    noisy_markers = ["烈士，男", "籍贯", "出生", "参加革命任", "时年"]
    return sum(1 for marker in noisy_markers if marker in text) >= 2


def normalize_zero_forms(value):
    return cell_text(value).replace("〇", "O").replace("○", "O").replace("0", "O").replace("零", "O")


def unit_role_change_is_worth_reporting(old_value, new_value):
    old = cell_text(old_value)
    new = cell_text(new_value)
    if not new or "/" not in new:
        return False
    if comparable(normalize_zero_forms(old)) == comparable(normalize_zero_forms(new)):
        return False
    if "/" in old and not old.endswith("/") and not re.search(r"(.+)/\1$", old):
        old_left, old_right = old.split("/", 1)
        new_left, new_right = new.split("/", 1)
        if old_left and old_right:
            if old_left.endswith(old_right) and new_right and comparable(old_right) == comparable(new_right):
                return True
            # Current v4 already has a usable unit/role separator. Do not
            # report content-only OCR disagreements as slash-format fixes.
            if old_right and new_right and comparable(old_right) != comparable(new_right):
                return False
            if comparable(old_left) != comparable(new_left):
                return False
    return True


def render_table(headers, rows, limit=None):
    if limit is not None:
        rows = rows[:limit]
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(md_escape(v) for v in row) + " |")
    return "\n".join(lines)


def render_markdown(report):
    s = report["summary"]
    lines = [
        "# 20260626 MinerU -> v5 Dry-run 核对报告",
        "",
        "## Summary",
        "",
        f"- 生成时间：`{report['generated_at']}`",
        f"- v4 文件：`{report['input']['v4_xlsx']}`",
        f"- 合并记录数：`{s['records_total']}`（预期 `243`）",
        f"- correction_items 总数：`{s['correction_item_total']}`",
        f"- 质量统计：`{json.dumps(s['quality_counts'], ensure_ascii=False)}`",
        f"- 告警统计：`{json.dumps(s['warning_counts'], ensure_ascii=False)}`",
        f"- Excel 未匹配行：`{s['rows_not_found']}`",
        f"- 姓名不一致：`{s['name_mismatches']}`",
        f"- 需复核记录：`{s['review_needed']}`",
        f"- 拟修改单元格：`{s['proposed_changes']}`",
        f"- 拟修改列统计：`{json.dumps(s['proposed_changes_by_column'], ensure_ascii=False)}`",
        f"- 无目标列解析项：`{s['unmapped_items']}`",
        f"- 跳过低可信候选：`{s['skipped_low_confidence_changes']}`",
        f"- T 列备份已填记录数：`{s['backup_populated_for_matched_records']}`",
        "",
        "说明：这是 dry-run，不写入 v4，也不生成 v5。J 列会按 `/` 拆分单位/职务；G/H 日期会去掉“年/月/日”；K 列只生成简要事迹候选。",
        "",
    ]

    if s["record_duplicate_codes"]:
        lines += ["## Duplicate Codes In Manifests", "", "```json", json.dumps(s["record_duplicate_codes"], ensure_ascii=False, indent=2), "```", ""]
    if s["excel_duplicate_codes"]:
        lines += ["## Duplicate Codes In Excel", "", "```json", json.dumps(s["excel_duplicate_codes"], ensure_ascii=False, indent=2), "```", ""]

    if report["rows_not_found"]:
        lines += ["## Rows Not Found", "", render_table(["code", "name", "source"], [[r["code"], r["name"], r["source_stem"]] for r in report["rows_not_found"]]), ""]

    if report["name_mismatches"]:
        lines += [
            "## Name Mismatches",
            "",
            render_table(["row", "code", "excel_name", "ocr_name"], [[r["row"], r["code"], r["excel_name"], r["ocr_name"]] for r in report["name_mismatches"]]),
            "",
        ]

    if report["review_needed"]:
        lines += [
            "## Review Needed",
            "",
            render_table(
                ["code", "name", "source", "quality", "warnings", "items", "markers"],
                [[r["code"], r["name"], r["source_stem"], r["quality"], ",".join(r["warnings"]), r["items"], r["markers"]] for r in report["review_needed"]],
            ),
            "",
        ]

    if report["unmapped_items"]:
        lines += [
            "## Unmapped Items",
            "",
            render_table(
                ["row", "code", "name", "field", "value"],
                [[r["row"], r["code"], r["name"], r["field"], r["value"]] for r in report["unmapped_items"]],
                limit=120,
            ),
            "",
            f"> 仅显示前 120 条；完整列表见 JSON。",
            "",
        ]

    if report["skipped_low_confidence_changes"]:
        lines += [
            "## Skipped Low-confidence Changes",
            "",
            render_table(
                ["row", "code", "name", "field", "old", "new", "reason"],
                [[r["row"], r["code"], r["name"], r["field"], r["old"], r["new"], r["reason"]] for r in report["skipped_low_confidence_changes"]],
            ),
            "",
        ]

    if report["proposed_changes"]:
        lines += [
            "## Proposed Changes",
            "",
            render_table(
                ["row", "code", "name", "col", "field", "old", "new"],
                [[c["row"], c["code"], c["name"], c["column"], c["field"], c["old"], c["new"]] for c in report["proposed_changes"]],
                limit=300,
            ),
            "",
            f"> 仅显示前 300 条；完整列表见 JSON。",
            "",
        ]

    return "\n".join(lines)


if __name__ == "__main__":
    main()
