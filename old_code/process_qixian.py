# -*- coding: utf-8 -*-
"""
祁县烈士英名录：PDF修正审核表 vs Excel 逐列比对审核
- 按编号匹配；编号不同则按姓名匹配
- 全字段扫描提取，不漏任何修正内容
- 籍贯：核心村名不同才算实质差异；符号差异跳过
- 其他字段：normalize后不同即修正，写红字+N列
"""
import json, re, os, shutil, openpyxl
from pathlib import Path
from openpyxl.styles import Font

LQ = '\u201c'
RQ = '\u201d'
RED = 'FFFF0000'

def n2(s):
    if s is None: return ''
    s = str(s).strip()
    s = s.replace('\\', '/')
    s = s.replace('\uff08','(').replace('\uff09',')')
    s = s.replace('（','(').replace('）',')')
    s = s.replace('\u3007','O')
    s = re.sub(r'[\s\n\r]+', '', s)
    return s.lower()

def set_red(cell):
    old = cell.font
    cell.font = Font(
        name=str(old.name) if old.name else 'Calibri',
        size=old.size if old.size else 11,
        bold=old.bold, italic=old.italic,
        color=RED
    )

def extract_all_corrections(corr_text):
    """用精确正则模式匹配各字段。

    格式1: 字段名补充完善为"值"理由：...（常见）
    格式2: 字段名补充完善为"值"（括号补充）理由：...（如"祁县梁坪寨民兵"（自卫队长））
    """
    if not corr_text: return {}
    items = {}
    patterns = [
        ('姓名',       r'姓名补充(?:填写|完善)为"([^"]+)"'),
        ('性别',       r'性别补充(?:填写|完善)为"([^"]+)"'),
        ('民族',       r'民族补充(?:填写|完善)为"([^"]+)"'),
        ('政治面貌',   r'政治面貌补充(?:填写|完善)为"([^"]+)"'),
        ('籍贯',       r'籍贯补充(?:填写|完善)为"([^"]+)"'),
        # 生前单位职务：引号后可能跟括号补充（如"祁县梁坪寨民兵"（自卫队长）），一直取到"理由"
        ('生前单位职务', r'生前（部队）单位及(?:曾任)?职务补充(?:填写|完善)为"([^"]*(?:"[^"]*)?)"'),
        ('出生时间',   r'出生(?:年月(?:日)?)?补充(?:填写|完善)为"([^"]+)"'),
        ('参加革命时间', r'参加革命[（工作]?(?:时间)?补充(?:填写|完善)为"([^"]+)"'),
        ('牺牲时间',   r'牺牲(?:年月(?:日)?)?补充(?:填写|完善)为"([^"]+)"'),
        ('牺牲地点',   r'牺牲地点补充(?:填写|完善)为"([^"]+)"'),
        ('牺牲原因',   r'牺牲原因补充(?:填写|完善)为"([^"]+)"'),
        ('事迹',       r'事迹补充(?:填写|完善)为"([^"]+)"'),
        ('出生时间',   r'出年时间补充(?:填写|完善)为"([^"]+)"'),
    ]
    for field, pat in patterns:
        m = re.search(pat, corr_text)
        if m:
            val = m.group(1).strip()
            # 如果引号后有括号补充（如"民兵"（队长）），补上括号内容
            # 查找引号结束位置，看后面是否有括号
            after_quote = corr_text[m.end(1):]
            bracket_m = re.match(r'（[^）]+）', after_quote)
            if bracket_m:
                val = val + bracket_m.group(0)
            # 同时处理引号内嵌套引号的情况（如"xxx"yyy"zzz"），展开
            if val.count('"') > 1:
                val = val.replace('"', '')
            items[field] = val
    return items

def get_shici_jjg(shici_val):
    """从事迹文本中提取籍贯"""
    if not shici_val: return None
    m = re.search(r'籍贯([^。，,]+)', str(shici_val))
    return m.group(1).strip() if m else None

def jjg_core(s):
    if not s: return ''
    s = str(s).replace('山西省晋中市祁县','').replace('山西省祁县','')
    s = re.sub(r'[\s\n\r\-/]+', '', s)
    return n2(s)

def is_real_jjg_diff(excel_val, pdf_val):
    """籍贯是否实质差异"""
    e = jjg_core(excel_val)
    p = jjg_core(pdf_val)
    if e and p and e != p: return True
    # 镇名比较
    def xz(s):
        for pat in [r'祁县([^\s区县]+镇)', r'祁县([^\s区县]+乡)', r'祁县([^\s区县]+公社)', r'([^\s区县]+镇)', r'([^\s区县]+乡)', r'([^\s区县]+公社)']:
            m = re.search(pat, str(s))
            if m: return n2(m.group(1))
        return ''
    if xz(excel_val) and xz(pdf_val) and xz(excel_val) != xz(pdf_val): return True
    return False

EXCEL_COL_MAP = {
    4: '姓名', 5: '性别', 6: '籍贯', 7: '出生时间',
    8: '参加革命时间', 9: '政治面貌', 10: '生前单位职务', 11: '牺牲时间地点',
}
FIELD_CN = {
    '姓名': '姓名', '性别': '性别', '籍贯': '籍贯',
    '出生时间': '出生时间', '参加革命时间': '参加革命时间',
    '政治面貌': '政治面貌', '生前单位职务': '生前单位职务',
    '牺牲时间': '牺牲时间', '牺牲地点': '牺牲地点',
    '牺牲原因': '牺牲原因', '事迹': '事迹',
}

# ── 加载Excel ────────────────────────────────────────────
EXCEL_SRC = r'd:\ying_min_mineru\《英名录》25版-排版-祁县-第一稿20260310B打印版-胡毅排3.13.xlsx'
OUT_DIR = r'd:\ying_min_mineru\output'
os.makedirs(OUT_DIR, exist_ok=True)
shutil.copy2(EXCEL_SRC, OUT_DIR + r'\英名录25版-祁县-二审_v3.xlsx')
excel_wb = openpyxl.load_workbook(OUT_DIR + r'\英名录25版-祁县-二审_v3.xlsx')
excel_ws = excel_wb.active

code_to_row = {}
name_to_row = {}
for row in range(2, excel_ws.max_row + 1):
    code = excel_ws.cell(row, 2).value
    name = excel_ws.cell(row, 4).value
    if code: code_to_row[str(code)] = row
    if name and str(name).strip(): name_to_row[str(name).strip()] = row

# ── 处理 ─────────────────────────────────────────────────
ext_out_dir = Path(r'd:\ying_min_mineru\extracted_out')
stats = {'total': 0, 'matched': 0, 'by_name': 0,
         'jjg_real': 0, 'jjg_format_skip': 0,
         'other_diff': 0, 'no_diff': 0}

for jf in sorted(ext_out_dir.glob('*.json')):
    stats['total'] += 1
    with open(jf, encoding='utf-8') as f:
        data = json.load(f)

    code = data.get('code', '')
    if not code: continue

    # 提取修正内容全文
    md = data.get('markdown', '')
    m = re.search(r'修正内容及理由</td><td[^>]*>(.*?)(?=</td></tr>)', md, re.DOTALL)
    corr_text = re.sub(r'<[^>]+>', '', m.group(1)).strip() if m else ''
    corrections = extract_all_corrections(corr_text)

    # 事迹文本中的籍贯（补充参考）
    shici_val = corrections.get('事迹', '')
    pdf_jjg = corrections.get('籍贯') or get_shici_jjg(shici_val)

    # 找Excel行
    excel_row = code_to_row.get(code)
    if excel_row is None:
        name = data.get('name', '').strip() or corrections.get('姓名', '')
        if name and name in name_to_row:
            excel_row = name_to_row[name]
            stats['by_name'] += 1
    if excel_row is None: continue

    stats['matched'] += 1
    row_notes = []
    real_diffs = []

    for col, field in EXCEL_COL_MAP.items():
        excel_val = excel_ws.cell(excel_row, col).value

        # 籍贯：特殊处理
        if field == '籍贯':
            if not pdf_jjg: continue
            if is_real_jjg_diff(excel_val, pdf_jjg):
                real_diffs.append((col, field, excel_val, pdf_jjg))
                row_notes.append('籍贯(应以PDF为准)')
                stats['jjg_real'] += 1
            else:
                stats['jjg_format_skip'] += 1
            continue

        # 其他字段
        pdf_val = corrections.get(field)
        if not pdf_val: continue

        e = n2(excel_val)
        p = n2(pdf_val)
        if e != p:
            real_diffs.append((col, field, excel_val, pdf_val))
            row_notes.append(FIELD_CN.get(field, field) + '(应以PDF为准)')
            stats['other_diff'] += 1

    # 写红字
    for col, field, old_val, new_val in real_diffs:
        cell = excel_ws.cell(excel_row, col)
        cell.value = new_val
        set_red(cell)

    # 写N列
    if row_notes:
        existing_n = excel_ws.cell(excel_row, 14).value
        note_str = '；'.join(row_notes)
        cell_n = excel_ws.cell(excel_row, 14)
        cell_n.value = (str(existing_n) + '\n' + note_str) if existing_n else note_str
        set_red(cell_n)
    elif not real_diffs:
        stats['no_diff'] += 1

# ── 保存 ─────────────────────────────────────────────────
out_path = OUT_DIR + r'\英名录25版-祁县-二审_v3.xlsx'
excel_wb.save(out_path)

print('处理完成！')
print('=' * 50)
print('总计JSON: ' + str(stats['total']))
print('匹配到Excel: ' + str(stats['matched']))
print('  按编号: ' + str(stats['matched'] - stats['by_name']))
print('  按姓名: ' + str(stats['by_name']))
print()
print('籍贯实质差异(红字+N列): ' + str(stats['jjg_real']))
print('籍贯格式差异(跳过): ' + str(stats['jjg_format_skip']))
print('其他字段差异(红字+N列): ' + str(stats['other_diff']))
print('无差异: ' + str(stats['no_diff']))
print()
print('输出: ' + out_path)
