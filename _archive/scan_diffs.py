# -*- coding: utf-8 -*-
"""
彻底扫描：哪些字段有内容差异（严格比对，不分格式/内容）
用于摸底，不写文件
"""
import json, re
from pathlib import Path

LQ = '\u201c'; RQ = '\u201d'

def n2(s):
    if s is None: return ''
    s = str(s).strip()
    s = s.replace('\\', '/')
    s = s.replace('\uff08','(').replace('\uff09',')')
    s = s.replace('（','(').replace('）',')')
    s = s.replace('\u3007','O')
    s = re.sub(r'[\s\n\r]+', '', s)
    return s.lower()

def extract_all_corrections(corr_text):
    """提取所有字段的修正值"""
    items = {}
    # 扫描所有 "XXX补充...为" 或 "XXX补充...为" 模式
    # 分两步：先定位字段名，再找对应的值
    field_names = ['姓名','性别','籍贯','出生时间','参加革命时间','政治面貌',
                   '生前单位职务','牺牲时间','牺牲地点','牺牲原因','事迹',
                   '民族','出生年月','牺牲年月']
    for fname in field_names:
        # 各种可能的格式：出生时间、出年、出生年月...
        for prefix in [fname, fname.replace('时间',''), fname.replace('年月','月')]:
            if not prefix: continue
            # Pattern: 字段名 + (可能的文字) + 补充 + (可能的文字) + 为 + "值"
            pat = re.escape(prefix) + r'(?:时间|年月|日期)?(?:补充(?:填写|完善)为|补充为|完善为)' + LQ + r'([^' + RQ + r']+)' + RQ
            m = re.search(pat, corr_text)
            if m:
                items[fname] = m.group(1).strip()
                break
    return items

def get_shici_jjg(shici_text):
    if not shici_text: return None
    m = re.search(r'籍贯([^。，,]+)', shici_text)
    return m.group(1).strip() if m else None

# 加载Excel
import openpyxl
wb = openpyxl.load_workbook(r'd:\ying_min_mineru\《英名录》25版-排版-祁县-第一稿20260310B打印版-胡毅排3.13.xlsx')
ws = wb.active
code_to_row = {}
name_to_row = {}
for row in range(2, ws.max_row+1):
    code = ws.cell(row,2).value
    name = ws.cell(row,4).value
    if code: code_to_row[str(code)] = row
    if name and str(name).strip(): name_to_row[str(name).strip()] = row

# 字段配置
FIELDS = {
    4: '姓名', 5: '性别', 6: '籍贯', 7: '出生时间',
    8: '参加革命时间', 9: '政治面貌', 10: '生前单位职务', 11: '牺牲时间地点'
}
FIELD_ALTS = {
    '籍贯': ['籍贯'],
    '出生时间': ['出生时间','出年时间','出生年月'],
    '政治面貌': ['政治面貌'],
    '生前单位职务': ['生前单位职务','生前单位'],
    '牺牲时间': ['牺牲时间'],
}

ext_out_dir = Path(r'd:\ying_min_mineru\extracted_out')

# 统计
stats = {}
for col, field in FIELDS.items():
    stats[field] = {'total_diff': 0, 'examples': []}

def get_corr_val(corrections, field):
    if field in corrections: return corrections[field]
    for alt in FIELD_ALTS.get(field, []):
        if alt in corrections: return corrections[alt]
    return None

for jf in sorted(ext_out_dir.glob('*.json')):
    with open(jf, encoding='utf-8') as f:
        data = json.load(f)

    code = data.get('code','')
    if not code: continue

    # 找Excel行
    excel_row = code_to_row.get(code)
    if excel_row is None:
        name = data.get('name','').strip()
        if name in name_to_row:
            excel_row = name_to_row[name]
    if excel_row is None: continue

    # 解析修正内容
    md = data.get('markdown','')
    m = re.search(r'修正内容及理由</td><td[^>]*>(.*?)(?=</td></tr>)', md, re.DOTALL)
    corr_text = re.sub(r'<[^>]+>','',m.group(1)).strip() if m else ''
    corrections = extract_all_corrections(corr_text)
    shici = corrections.get('事迹','')
    pdf_jjg = corrections.get('籍贯') or get_shici_jjg(shici)

    for col, field in FIELDS.items():
        excel_val = ws.cell(excel_row, col).value
        # 取PDF修正值
        if field == '籍贯':
            pdf_val = pdf_jjg
        else:
            pdf_val = get_corr_val(corrections, field)

        if not pdf_val: continue
        e = n2(excel_val)
        p = n2(pdf_val)
        if e != p:
            stats[field]['total_diff'] += 1
            if len(stats[field]['examples']) < 3:
                stats[field]['examples'].append({
                    'code': code, 'excel': excel_val, 'pdf': pdf_val
                })

print('=== 各字段内容差异统计 ===')
for field, info in stats.items():
    print(field + ': ' + str(info['total_diff']) + '处')
    for ex in info['examples']:
        print('  [' + ex['code'] + ']')
        print('    Excel: ' + repr(str(ex['excel'])[:50] if ex['excel'] else None))
        print('    PDF:   ' + repr(str(ex['pdf'])[:50]))

print()
print('=== 籍贯差异详细（全部）===')
for jf in sorted(ext_out_dir.glob('*.json')):
    with open(jf, encoding='utf-8') as f:
        data = json.load(f)
    code = data.get('code','')
    if not code: continue
    excel_row = code_to_row.get(code)
    if excel_row is None:
        name = data.get('name','').strip()
        if name in name_to_row: excel_row = name_to_row[name]
    if excel_row is None: continue

    md = data.get('markdown','')
    m = re.search(r'修正内容及理由</td><td[^>]*>(.*?)(?=</td></tr>)', md, re.DOTALL)
    corr_text = re.sub(r'<[^>]+>','',m.group(1)).strip() if m else ''
    corrections = extract_all_corrections(corr_text)
    shici = corrections.get('事迹','')
    pdf_jjg = corrections.get('籍贯') or get_shici_jjg(shici)
    excel_jjg = ws.cell(excel_row,6).value
    if pdf_jjg and n2(excel_jjg) != n2(pdf_jjg):
        # 排除纯格式差异（去除省市区镇后村名相同）
        def core(s):
            s = str(s).replace('山西省晋中市祁县','').replace('山西省祁县','')
            s = re.sub(r'[\s\n\r\-/]+','',s)
            return n2(s)
        e_c = core(excel_jjg)
        p_c = core(pdf_jjg)
        diff_type = '内容' if e_c != p_c else '格式'
        print('[' + code + '] (' + diff_type + ')')
        print('  Excel: ' + repr(excel_jjg))
        print('  PDF:   ' + repr(pdf_jjg))
