"""Build a v6-specific dry-run from explicit correction items."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
import json
from pathlib import Path
import re
from typing import Any, Callable

from openpyxl.utils import get_column_letter

from .index import index_rows_by_code, load_workbook_sheet
from .rules import cell_text, comparable
from .v6_rules import format_v6_origin, format_v6_story, format_v6_unit_role, normalize_v6_date

JsonDict = dict[str, Any]

COL_SEX = 5
COL_ORIGIN = 6
COL_BIRTH = 7
COL_JOIN = 8
COL_POLITICAL = 9
COL_UNIT_ROLE = 10
COL_STORY = 11

_DUPLICATE_DELETE = re.compile(r"(?:重复|与编号.*复).*(?:删除|册.{0,2}除|刑除|删|除)")


def build_v6_dry_run(base_xlsx: str | Path, records_path: str | Path) -> JsonDict:
    """Return safe v6 changes while leaving N and T untouched."""

    base = Path(base_xlsx)
    records_file = Path(records_path)
    payload = json.loads(records_file.read_text(encoding="utf-8"))
    records = payload.get("records", [])
    if not isinstance(records, list):
        raise ValueError("records must be a list")

    workbook, sheet = load_workbook_sheet(base)
    try:
        row_index = index_rows_by_code(sheet)
        duplicate_excel_codes = _duplicate_excel_codes(sheet)
        proposed: list[JsonDict] = []
        blocked: list[JsonDict] = []
        unmapped: list[JsonDict] = []
        record_results: list[JsonDict] = []

        for record in records:
            if not isinstance(record, dict):
                continue
            result = _build_record_result(
                record=record,
                sheet=sheet,
                row_index=row_index,
                duplicate_excel_codes=duplicate_excel_codes,
                proposed=proposed,
                blocked=blocked,
                unmapped=unmapped,
            )
            record_results.append(result)

        counts = Counter(change["column_letter"] for change in proposed)
        status_counts = Counter(result["status"] for result in record_results)
        unmapped_counts = Counter(item["field"] or item["raw_label"] or "<unrecognized>" for item in unmapped)
        return {
            "profile": "v6_candidate",
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "base_xlsx": base.as_posix(),
            "records_path": records_file.as_posix(),
            "rules": {
                "trusted_baseline": "v4",
                "date_format": "YYYY-M or YYYY-M-D; keep year-only and 0",
                "long_text_red": "changed_runs_only",
                "short_field_red": "whole_cell",
                "story_source": "explicit structured 事迹 item only",
                "review_column_n": "preserved",
                "backup_column_t": "preserved",
            },
            "summary": {
                "record_count": len(record_results),
                "status_counts": dict(sorted(status_counts.items())),
                "proposed_change_count": len(proposed),
                "proposed_changes_by_column": dict(sorted(counts.items())),
                "blocked_change_count": len(blocked),
                "unmapped_item_count": len(unmapped),
                "unmapped_items_by_field": dict(sorted(unmapped_counts.items())),
                "duplicate_excel_codes": duplicate_excel_codes,
            },
            "proposed_changes": proposed,
            "blocked_changes": blocked,
            "unmapped_items": unmapped,
            "record_results": record_results,
        }
    finally:
        workbook.close()


def write_v6_dry_run(report: JsonDict, json_path: str | Path, markdown_path: str | Path) -> None:
    json_target = Path(json_path)
    markdown_target = Path(markdown_path)
    json_target.parent.mkdir(parents=True, exist_ok=True)
    markdown_target.parent.mkdir(parents=True, exist_ok=True)
    json_target.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    markdown_target.write_text(render_v6_dry_run_markdown(report), encoding="utf-8")


def render_v6_dry_run_markdown(report: JsonDict) -> str:
    summary = report["summary"]
    lines = [
        "# v6 Candidate Dry-run",
        "",
        f"- base: `{report['base_xlsx']}`",
        f"- records: `{report['records_path']}`",
        f"- records processed: {summary['record_count']}",
        f"- status counts: `{json.dumps(summary['status_counts'], ensure_ascii=False)}`",
        f"- proposed changes: {summary['proposed_change_count']}",
        f"- changes by column: `{json.dumps(summary['proposed_changes_by_column'], ensure_ascii=False)}`",
        f"- blocked changes: {summary['blocked_change_count']}",
        f"- unmapped correction items: {summary['unmapped_item_count']}",
        "",
        "N and T are preserved. K uses only the complete explicit `事迹` correction item.",
        "",
        "## Blocked Changes",
        "",
        "| Code | Name | Field | Old | Candidate | Reason |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for item in report["blocked_changes"]:
        lines.append(
            f"| {_table(item.get('code', ''))} | {_table(item.get('name', ''))} | "
            f"{_table(item.get('field', ''))} | {_table(item.get('old', ''))} | "
            f"{_table(item.get('new', ''))} | {_table(item.get('block_reason', ''))} |"
        )
    lines.extend(["", "## Record Status", ""])
    for item in report["record_results"]:
        if item["status"] != "matched":
            lines.append(
                f"- {item.get('code') or 'MISSING_CODE'} / {item.get('name') or 'MISSING_NAME'}: "
                f"{item['status']} ({', '.join(item.get('warnings', [])) or 'no warning'})"
            )
    return "\n".join(lines).rstrip() + "\n"


def _build_record_result(
    record: JsonDict,
    sheet: Any,
    row_index: dict[str, Any],
    duplicate_excel_codes: list[str],
    proposed: list[JsonDict],
    blocked: list[JsonDict],
    unmapped: list[JsonDict],
) -> JsonDict:
    code = cell_text(record.get("code"))
    name = cell_text(record.get("name"))
    source_stem = cell_text(record.get("source_stem"))
    raw_text = cell_text(record.get("raw_text"))
    warnings = [str(value) for value in record.get("warnings", []) if isinstance(value, str)]
    result = {
        "code": code,
        "name": name,
        "source_stem": source_stem,
        "row": 0,
        "status": "matched",
        "warnings": list(warnings),
        "change_ids": [],
        "blocked_change_ids": [],
    }

    if not code:
        result["status"] = "code_missing"
        return result
    if code in duplicate_excel_codes:
        result["status"] = "duplicate_code_in_v4"
        return result
    indexed = row_index.get(code)
    if indexed is None:
        if _DUPLICATE_DELETE.search(raw_text):
            result["status"] = "duplicate_delete_already_absent"
        else:
            result["status"] = "code_not_found"
        return result
    result["row"] = indexed.row
    if not name or comparable(indexed.name) != comparable(name):
        result["status"] = "name_mismatch"
        result["warnings"].append(f"expected_name:{indexed.name}")
        return result

    items = [item for item in record.get("items", []) if isinstance(item, dict)]
    by_field: dict[str, list[JsonDict]] = defaultdict(list)
    for item in items:
        field_name = cell_text(item.get("field"))
        if field_name:
            by_field[field_name].append(item)
        if field_name not in {
            "性别",
            "籍贯",
            "出生时间",
            "参加革命/工作时间",
            "政治面貌",
            "生前单位及曾任职务",
            "曾任职务",
            "事迹",
        }:
            unmapped.append(
                {
                    "code": code,
                    "name": name,
                    "source_stem": source_stem,
                    "field": field_name,
                    "raw_label": cell_text(item.get("raw_label")),
                    "value": cell_text(item.get("value")),
                    "reason": cell_text(item.get("reason")),
                    "warnings": _string_list(item.get("warnings")),
                }
            )

    specs: list[tuple[tuple[str, ...], int, str, Callable[[object], str | None], str]] = [
        (("性别",), COL_SEX, "性别", lambda value: cell_text(value), "full"),
        (("籍贯",), COL_ORIGIN, "籍贯", format_v6_origin, "partial"),
        (("出生时间",), COL_BIRTH, "出生时间", normalize_v6_date, "full"),
        (("参加革命/工作时间",), COL_JOIN, "参加革命/工作时间", normalize_v6_date, "full"),
        (("政治面貌",), COL_POLITICAL, "政治面貌", lambda value: cell_text(value), "full"),
        (("生前单位及曾任职务", "曾任职务"), COL_UNIT_ROLE, "生前单位及曾任职务", format_v6_unit_role, "partial"),
        (("事迹",), COL_STORY, "牺牲时间、地点及简要事迹", format_v6_story, "partial"),
    ]
    for field_names, column, target_label, formatter, red_mode in specs:
        candidates = [item for field_name in field_names for item in by_field.get(field_name, [])]
        if not candidates:
            continue
        prepared: list[tuple[str, JsonDict]] = []
        for item in candidates:
            raw_value = cell_text(item.get("value"))
            formatted = formatter(raw_value)
            if formatted is None:
                blocked_item = _blocked_item(
                    code, name, indexed.row, column, target_label, sheet.cell(indexed.row, column).value,
                    raw_value, "date_format_unrecognized", item,
                )
                blocked.append(blocked_item)
                result["blocked_change_ids"].append(blocked_item["id"])
                continue
            if formatted:
                prepared.append((formatted, item))
        unique_values = _unique_values(value for value, _item in prepared)
        if len(unique_values) > 1:
            blocked_item = _blocked_item(
                code, name, indexed.row, column, target_label, sheet.cell(indexed.row, column).value,
                " | ".join(unique_values), "multiple_correction_values", prepared[0][1],
            )
            blocked.append(blocked_item)
            result["blocked_change_ids"].append(blocked_item["id"])
            continue
        if not unique_values:
            continue
        new_value = unique_values[0]
        old_value = cell_text(sheet.cell(indexed.row, column).value)
        if comparable(old_value) == comparable(new_value):
            continue
        source_item = next(item for value, item in prepared if comparable(value) == comparable(new_value))
        change = {
            "id": f"{code}:{get_column_letter(column)}{indexed.row}",
            "code": code,
            "name": name,
            "source_stem": source_stem,
            "row": indexed.row,
            "column": column,
            "column_letter": get_column_letter(column),
            "field": target_label,
            "old": old_value,
            "new": new_value,
            "reason": cell_text(source_item.get("reason")),
            "warnings": _string_list(source_item.get("warnings")),
            "requires": [],
            "story_length": len(new_value) if column == COL_STORY else 0,
            "red_mode": red_mode,
        }
        proposed.append(change)
        result["change_ids"].append(change["id"])
    return result


def _blocked_item(
    code: str,
    name: str,
    row: int,
    column: int,
    field_name: str,
    old_value: object,
    new_value: object,
    reason: str,
    source_item: JsonDict,
) -> JsonDict:
    return {
        "id": f"{code}:{get_column_letter(column)}{row}",
        "code": code,
        "name": name,
        "row": row,
        "column": column,
        "column_letter": get_column_letter(column),
        "field": field_name,
        "old": cell_text(old_value),
        "new": cell_text(new_value),
        "block_reason": reason,
        "source_reason": cell_text(source_item.get("reason")),
        "warnings": _string_list(source_item.get("warnings")),
    }


def _duplicate_excel_codes(sheet: Any) -> list[str]:
    counts: Counter[str] = Counter()
    for row in range(2, sheet.max_row + 1):
        code = cell_text(sheet.cell(row, 2).value)
        if code:
            counts[code] += 1
    return sorted(code for code, count in counts.items() if count > 1)


def _unique_values(values: Any) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = comparable(value)
        if key and key not in seen:
            seen.add(key)
            result.append(value)
    return result


def _string_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value if isinstance(item, str)]


def _table(value: object) -> str:
    return cell_text(value).replace("|", "\\|").replace("\n", "<br>")
