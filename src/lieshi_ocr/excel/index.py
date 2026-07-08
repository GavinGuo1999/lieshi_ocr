"""Read-only Excel row indexing by martyr code."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .rules import COL_CODE, COL_NAME, cell_text


@dataclass(frozen=True)
class ExcelRow:
    code: str
    name: str
    row: int


def load_workbook_sheet(path: str | Path) -> tuple[openpyxl.Workbook, Worksheet]:
    workbook = openpyxl.load_workbook(path)
    return workbook, workbook.active


def index_rows_by_code(sheet: Worksheet, start_row: int = 2) -> dict[str, ExcelRow]:
    index: dict[str, ExcelRow] = {}
    for row in range(start_row, sheet.max_row + 1):
        code = cell_text(sheet.cell(row, COL_CODE).value)
        if not code:
            continue
        index[code] = ExcelRow(code=code, name=cell_text(sheet.cell(row, COL_NAME).value), row=row)
    return index
