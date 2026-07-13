"""Generate Excel update dry-run reports without modifying workbooks."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import json
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from .index import index_rows_by_code, load_workbook_sheet
from .rules import (
    BACKUP_CONFLICT,
    BACKUP_REQUIRED,
    COL_NAME,
    COL_REVIEW,
    COL_STORY,
    COL_STORY_BACKUP,
    COLUMN_LABELS,
    TARGET_FIELDS,
    STORY_REVIEW_LENGTH,
    append_review_note,
    build_review_note,
    build_story,
    cell_text,
    classify_story_backup,
    comparable,
    normalize_unit_role,
)

JsonDict = dict[str, Any]


@dataclass(frozen=True)
class ExcelChange:
    id: str
    code: str
    name: str
    row: int
    column: int
    column_letter: str
    field: str
    old: str
    new: str
    reason: str
    warnings: list[str] = field(default_factory=list)
    requires: list[str] = field(default_factory=list)
    story_length: int = 0

    def to_json(self) -> JsonDict:
        return {
            "id": self.id,
            "code": self.code,
            "name": self.name,
            "row": self.row,
            "column": self.column,
            "column_letter": self.column_letter,
            "field": self.field,
            "old": self.old,
            "new": self.new,
            "reason": self.reason,
            "warnings": self.warnings,
            "requires": self.requires,
            "story_length": self.story_length,
        }


@dataclass(frozen=True)
class BlockedChange:
    id: str
    code: str
    name: str
    row: int
    column: int
    column_letter: str
    field: str
    old: str
    new: str
    reason: str
    block_reason: str
    warnings: list[str] = field(default_factory=list)
    story_length: int = 0

    def to_json(self) -> JsonDict:
        return {
            "id": self.id,
            "code": self.code,
            "name": self.name,
            "row": self.row,
            "column": self.column,
            "column_letter": self.column_letter,
            "field": self.field,
            "old": self.old,
            "new": self.new,
            "reason": self.reason,
            "block_reason": self.block_reason,
            "warnings": self.warnings,
            "story_length": self.story_length,
        }


@dataclass(frozen=True)
class RecordResult:
    code: str
    name: str
    source_stem: str
    row: int = 0
    warnings: list[str] = field(default_factory=list)
    change_ids: list[str] = field(default_factory=list)
    blocked_change_ids: list[str] = field(default_factory=list)

    def to_json(self) -> JsonDict:
        return {
            "code": self.code,
            "name": self.name,
            "source_stem": self.source_stem,
            "row": self.row,
            "warnings": self.warnings,
            "change_ids": self.change_ids,
            "blocked_change_ids": self.blocked_change_ids,
        }


@dataclass(frozen=True)
class ExcelDryRunReport:
    base_xlsx: str
    records_path: str
    generated_at: str
    proposed_changes: list[ExcelChange]
    blocked_changes: list[BlockedChange]
    record_results: list[RecordResult]
    warnings: list[str] = field(default_factory=list)

    def to_json(self) -> JsonDict:
        return {
            "base_xlsx": self.base_xlsx,
            "records_path": self.records_path,
            "generated_at": self.generated_at,
            "change_count": len(self.proposed_changes),
            "blocked_count": len(self.blocked_changes),
            "record_count": len(self.record_results),
            "proposed_changes": [change.to_json() for change in self.proposed_changes],
            "blocked_changes": [change.to_json() for change in self.blocked_changes],
            "record_results": [result.to_json() for result in self.record_results],
            "warnings": self.warnings,
        }


def build_excel_dry_run(base_xlsx: str | Path, records_path: str | Path) -> ExcelDryRunReport:
    """Build proposed Excel changes from correction_records.json."""

    workbook, sheet = load_workbook_sheet(base_xlsx)
    try:
        row_index = index_rows_by_code(sheet)
        records_payload = json.loads(Path(records_path).read_text(encoding="utf-8"))
        records = records_payload.get("records", [])
        proposed: list[ExcelChange] = []
        blocked: list[BlockedChange] = []
        results: list[RecordResult] = []
        warnings: list[str] = []
        if not isinstance(records, list):
            records = []
            warnings.append("records_not_list")

        for record in records:
            if not isinstance(record, dict):
                warnings.append("record_not_object")
                continue
            result, changes, blocked_changes = _build_record_changes(record, sheet, row_index)
            results.append(result)
            proposed.extend(changes)
            blocked.extend(blocked_changes)

        return ExcelDryRunReport(
            base_xlsx=Path(base_xlsx).as_posix(),
            records_path=Path(records_path).as_posix(),
            generated_at=datetime.now().isoformat(timespec="seconds"),
            proposed_changes=proposed,
            blocked_changes=blocked,
            record_results=results,
            warnings=warnings,
        )
    finally:
        workbook.close()


def write_dry_run_outputs(report: ExcelDryRunReport, json_path: str | Path, markdown_path: str | Path) -> None:
    json_target = Path(json_path)
    md_target = Path(markdown_path)
    json_target.parent.mkdir(parents=True, exist_ok=True)
    md_target.parent.mkdir(parents=True, exist_ok=True)
    json_target.write_text(json.dumps(report.to_json(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    md_target.write_text(render_dry_run_markdown(report), encoding="utf-8")


def render_dry_run_markdown(report: ExcelDryRunReport) -> str:
    lines = [
        "# Excel Dry Run Report",
        "",
        f"- base_xlsx: `{report.base_xlsx}`",
        f"- records_path: `{report.records_path}`",
        f"- proposed changes: {len(report.proposed_changes)}",
        f"- blocked changes: {len(report.blocked_changes)}",
        "",
        "| ID | Code | Name | Cell | Field | Old | New | Story Length | Requires |",
        "| --- | --- | --- | --- | --- | --- | --- | ---: | --- |",
    ]
    for change in report.proposed_changes:
        cell = f"{change.column_letter}{change.row}"
        lines.append(
            f"| {change.id} | {change.code} | {change.name} | {cell} | {change.field} | "
            f"{_table(change.old)} | {_table(change.new)} | {change.story_length or ''} | "
            f"{', '.join(change.requires) or ''} |"
        )
    lines.extend(
        [
            "",
            "## Blocked Changes",
            "",
            "| ID | Code | Name | Cell | Old | Candidate | Block Reason | Story Length | Warnings |",
            "| --- | --- | --- | --- | --- | --- | --- | ---: | --- |",
        ]
    )
    for change in report.blocked_changes:
        cell = f"{change.column_letter}{change.row}"
        lines.append(
            f"| {change.id} | {change.code} | {change.name} | {cell} | {_table(change.old)} | "
            f"{_table(change.new)} | {change.block_reason} | {change.story_length} | "
            f"{', '.join(change.warnings)} |"
        )
    lines.extend(["", "## Record Warnings", ""])
    for result in report.record_results:
        if result.warnings:
            lines.append(f"- {result.code or 'MISSING_CODE'} / {result.name or 'MISSING_NAME'}: {', '.join(result.warnings)}")
    return "\n".join(lines).rstrip() + "\n"


def _build_record_changes(
    record: JsonDict,
    sheet: Any,
    row_index: dict[str, Any],
) -> tuple[RecordResult, list[ExcelChange], list[BlockedChange]]:
    code = cell_text(record.get("code"))
    name = cell_text(record.get("name"))
    source_stem = cell_text(record.get("source_stem"))
    fields = record.get("fields", {})
    if not isinstance(fields, dict):
        fields = {}
    record_warnings = [str(item) for item in record.get("warnings", []) if isinstance(item, str)]
    warnings = list(record_warnings)

    if not code:
        warnings.append("code_missing")
        return RecordResult(code=code, name=name, source_stem=source_stem, warnings=_dedupe(warnings)), [], []
    indexed = row_index.get(code)
    if indexed is None:
        warnings.append("code_not_found")
        return RecordResult(code=code, name=name, source_stem=source_stem, warnings=_dedupe(warnings)), [], []
    if name and comparable(indexed.name) != comparable(name):
        warnings.append("name_mismatch")
        return RecordResult(code=code, name=name, source_stem=source_stem, row=indexed.row, warnings=_dedupe(warnings)), [], []

    normalized_fields = {str(key): cell_text(value) for key, value in fields.items()}
    story = build_story(normalized_fields)
    old_story = cell_text(sheet.cell(indexed.row, COL_STORY).value)
    existing_backup = cell_text(sheet.cell(indexed.row, COL_STORY_BACKUP).value)
    story_would_change = bool(story and comparable(old_story) != comparable(story))
    backup_status = classify_story_backup(old_story, existing_backup) if story_would_change else ""
    story_length = len(story)
    if story_would_change and story_length > STORY_REVIEW_LENGTH:
        warnings.append("story_candidate_long")
    if story_would_change and backup_status == BACKUP_CONFLICT:
        warnings.append("story_backup_conflict")
    warnings = _dedupe(warnings)

    changes: list[ExcelChange] = []
    blocked_changes: list[BlockedChange] = []
    for field_name, column in TARGET_FIELDS.items():
        new_value = cell_text(fields.get(field_name))
        if field_name == "生前单位及曾任职务":
            new_value = normalize_unit_role(new_value, cell_text(fields.get("曾任职务")))
        _append_change(changes, sheet, indexed.row, column, code, name or indexed.name, new_value, f"{field_name} from review record", warnings)

    if story_would_change and backup_status == BACKUP_CONFLICT:
        blocked_changes.append(
            _blocked_story_change(
                row=indexed.row,
                code=code,
                name=name or indexed.name,
                old_story=old_story,
                story=story,
                warnings=warnings,
            )
        )
    elif story_would_change:
        requires: list[str] = []
        if backup_status == BACKUP_REQUIRED:
            backup_change = _append_change(
                changes,
                sheet,
                indexed.row,
                COL_STORY_BACKUP,
                code,
                name or indexed.name,
                old_story,
                "backup old K before story update",
                warnings,
            )
            if backup_change is None:
                raise RuntimeError("Expected a T backup change before K update")
            requires.append(backup_change.id)
        _append_change(
            changes,
            sheet,
            indexed.row,
            COL_STORY,
            code,
            name or indexed.name,
            story,
            "short sacrifice story",
            warnings,
            requires=requires,
            story_length=story_length,
        )

    safe_business_changes = [
        change for change in changes if change.column not in {COL_REVIEW, COL_STORY_BACKUP}
    ]
    has_non_column_review_value = bool(cell_text(fields.get("安葬地")) or cell_text(fields.get("民族")))
    review_note = build_review_note(warnings, normalized_fields)
    existing_review_note = cell_text(sheet.cell(indexed.row, COL_REVIEW).value)
    appended_review_note = append_review_note(existing_review_note, review_note)
    if safe_business_changes or has_non_column_review_value:
        _append_change(
            changes,
            sheet,
            indexed.row,
            COL_REVIEW,
            code,
            name or indexed.name,
            appended_review_note,
            "append review warnings and non-column fields",
            warnings,
        )
    return (
        RecordResult(
            code=code,
            name=name or indexed.name,
            source_stem=source_stem,
            row=indexed.row,
            warnings=_dedupe(warnings),
            change_ids=[change.id for change in changes],
            blocked_change_ids=[change.id for change in blocked_changes],
        ),
        changes,
        blocked_changes,
    )


def _append_change(
    changes: list[ExcelChange],
    sheet: Any,
    row: int,
    column: int,
    code: str,
    name: str,
    new_value: str,
    reason: str,
    warnings: list[str],
    requires: list[str] | None = None,
    story_length: int = 0,
) -> ExcelChange | None:
    if not new_value:
        return None
    old_value = cell_text(sheet.cell(row, column).value)
    if comparable(old_value) == comparable(new_value):
        return None
    column_letter = get_column_letter(column)
    change = ExcelChange(
        id=f"{code}:{column_letter}{row}",
        code=code,
        name=name,
        row=row,
        column=column,
        column_letter=column_letter,
        field=COLUMN_LABELS[column],
        old=old_value,
        new=new_value,
        reason=reason,
        warnings=list(warnings),
        requires=list(requires or []),
        story_length=story_length,
    )
    changes.append(change)
    return change


def _blocked_story_change(
    row: int,
    code: str,
    name: str,
    old_story: str,
    story: str,
    warnings: list[str],
) -> BlockedChange:
    column_letter = get_column_letter(COL_STORY)
    return BlockedChange(
        id=f"{code}:{column_letter}{row}",
        code=code,
        name=name,
        row=row,
        column=COL_STORY,
        column_letter=column_letter,
        field=COLUMN_LABELS[COL_STORY],
        old=old_story,
        new=story,
        reason="short sacrifice story",
        block_reason="story_backup_conflict",
        warnings=list(warnings),
        story_length=len(story),
    )


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _table(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")
