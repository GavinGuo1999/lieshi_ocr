"""Generate and collect a protected Excel workbook for story review."""

from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from lieshi_ocr.excel.rules import cell_text
from lieshi_ocr.pipeline.audit_story import build_story_candidate_audit

JsonDict = dict[str, Any]

REVIEW_SHEET = "人工审核"
SOURCE_SHEET = "原始正文"
INSTRUCTIONS_SHEET = "使用说明"
METADATA_SHEET = "_审核元数据"

REVIEW_HEADERS = [
    "序号",
    "编号",
    "姓名",
    "当前K",
    "当前T",
    "系统候选K",
    "候选字数",
    "日期候选",
    "牺牲地点候选",
    "牺牲原因候选",
    "风险提示",
    "blocked/proposed状态",
    "审核结论",
    "建议最终K",
    "人工备注",
    "MinerU原文路径",
    "correction crop路径",
]

SOURCE_HEADERS = ["编号", "姓名", "完整MinerU正文", "来源文件", "warnings"]
METADATA_HEADERS = ["审核表行号", "编号", "姓名", "系统候选K", "候选状态", "阻断原因", "source_stem"]

REVIEW_CONCLUSIONS = [
    "通过系统候选",
    "需要改写",
    "信息不足",
    "原文疑似错误",
    "暂不处理",
]

PASS_CONCLUSION = REVIEW_CONCLUSIONS[0]
REWRITE_CONCLUSION = REVIEW_CONCLUSIONS[1]
BLOCKING_CONCLUSIONS = set(REVIEW_CONCLUSIONS[2:])

HEADER_FILL = PatternFill("solid", fgColor="264653")
LOCKED_FILL = PatternFill("solid", fgColor="F2F3F5")
EDITABLE_FILL = PatternFill("solid", fgColor="E2F0D9")
BLOCKED_FILL = PatternFill("solid", fgColor="FCE8E6")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Microsoft YaHei", color="202124", size=10)
THIN_GREY = Side(style="thin", color="D9DEE3")
CELL_BORDER = Border(bottom=THIN_GREY)
EDITABLE_BORDER = Border(left=THIN_GREY, right=THIN_GREY, bottom=THIN_GREY)


def generate_story_review_workbook(
    base_xlsx: str | Path,
    records_path: str | Path,
    dry_run_path: str | Path,
    out_xlsx: str | Path,
) -> JsonDict:
    """Build a review workbook from the existing read-only story audit inputs."""

    base = Path(base_xlsx)
    output = Path(out_xlsx)
    if base.resolve(strict=False) == output.resolve(strict=False):
        raise ValueError("Review workbook must not overwrite the trusted baseline")
    report = build_story_candidate_audit(records_path, dry_run_path, base)
    return write_story_review_workbook(report, output)


def write_story_review_workbook(report: JsonDict, out_xlsx: str | Path) -> JsonDict:
    records = report.get("records", [])
    if not isinstance(records, list):
        raise ValueError("story audit report field 'records' must be a list")

    workbook = Workbook()
    review = workbook.active
    review.title = REVIEW_SHEET
    source = workbook.create_sheet(SOURCE_SHEET)
    instructions = workbook.create_sheet(INSTRUCTIONS_SHEET)
    metadata = workbook.create_sheet(METADATA_SHEET)

    _write_review_sheet(review, records)
    _write_source_sheet(source, records)
    _write_instructions_sheet(instructions)
    _write_metadata_sheet(metadata, records)
    metadata.sheet_state = "veryHidden"

    workbook.properties.title = "事迹候选人工审核"
    workbook.properties.subject = "只读候选对比与人工审核，不修改可信 v4"
    output = Path(out_xlsx)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return {
        "out_xlsx": output.as_posix(),
        "record_count": len(records),
        "editable_columns": ["审核结论", "建议最终K", "人工备注"],
    }


def collect_story_review_decisions(review_xlsx: str | Path) -> JsonDict:
    """Read validated human decisions without modifying the review workbook."""

    source = Path(review_xlsx)
    workbook = openpyxl.load_workbook(source, data_only=False, read_only=False)
    try:
        required = {REVIEW_SHEET, METADATA_SHEET}
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"Review workbook missing sheets: {', '.join(sorted(missing))}")
        review = workbook[REVIEW_SHEET]
        metadata = workbook[METADATA_SHEET]
        _validate_headers(review, REVIEW_HEADERS)
        _validate_headers(metadata, METADATA_HEADERS)

        decisions: list[JsonDict] = []
        total_count = 0
        for meta_row in range(2, metadata.max_row + 1):
            review_row = _integer(metadata.cell(meta_row, 1).value)
            if not review_row:
                continue
            total_count += 1
            expected_code = cell_text(metadata.cell(meta_row, 2).value)
            expected_name = cell_text(metadata.cell(meta_row, 3).value)
            expected_candidate = cell_text(metadata.cell(meta_row, 4).value)
            expected_status = cell_text(metadata.cell(meta_row, 5).value)
            expected_block_reason = cell_text(metadata.cell(meta_row, 6).value)
            actual_code = cell_text(review.cell(review_row, 2).value)
            actual_name = cell_text(review.cell(review_row, 3).value)
            actual_candidate = cell_text(review.cell(review_row, 6).value)
            actual_status = cell_text(review.cell(review_row, 12).value)
            if actual_code != expected_code or actual_name != expected_name:
                raise ValueError(f"Row {review_row}: code or name differs from protected metadata")
            if actual_candidate != expected_candidate:
                raise ValueError(f"Row {review_row}: system story candidate differs from protected metadata")
            if actual_status != expected_status:
                raise ValueError(f"Row {review_row}: candidate status differs from protected metadata")

            conclusion = cell_text(review.cell(review_row, 13).value)
            suggested = cell_text(review.cell(review_row, 14).value)
            notes = cell_text(review.cell(review_row, 15).value)
            if not conclusion:
                continue
            if conclusion not in REVIEW_CONCLUSIONS:
                raise ValueError(f"Row {review_row}: unsupported review conclusion: {conclusion}")
            if conclusion == PASS_CONCLUSION:
                if not expected_candidate:
                    raise ValueError(f"Row {review_row}: system story candidate is empty")
                final_story = expected_candidate
                approvable = True
            elif conclusion == REWRITE_CONCLUSION:
                if not suggested:
                    raise ValueError(f"Row {review_row}: rewritten final story is required")
                final_story = suggested
                approvable = True
            else:
                final_story = ""
                approvable = False
            decisions.append(
                {
                    "code": expected_code,
                    "name": expected_name,
                    "review_conclusion": conclusion,
                    "suggested_final_story": suggested,
                    "notes": notes,
                    "final_story": final_story,
                    "content_approved": approvable,
                    "source_candidate_status": expected_status,
                    "source_block_reason": expected_block_reason,
                    "requires_backup_resolution": expected_status == "blocked",
                    "approvable": approvable and expected_status == "proposed",
                    "review_row": review_row,
                }
            )
    finally:
        workbook.close()

    passed_count = sum(1 for item in decisions if item["review_conclusion"] == PASS_CONCLUSION)
    rewritten_count = sum(1 for item in decisions if item["review_conclusion"] == REWRITE_CONCLUSION)
    blocked_count = sum(1 for item in decisions if item["review_conclusion"] in BLOCKING_CONCLUSIONS)
    return {
        "schema": "lieshi_ocr_story_review_decisions",
        "version": 1,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_review_xlsx": source.as_posix(),
        "summary": {
            "total_count": total_count,
            "reviewed_count": len(decisions),
            "passed_count": passed_count,
            "rewritten_count": rewritten_count,
            "blocked_count": blocked_count,
            "pending_count": total_count - len(decisions),
        },
        "decisions": decisions,
    }


def write_story_review_decisions(report: JsonDict, out_json: str | Path) -> Path:
    output = Path(out_json)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return output


def _write_review_sheet(sheet: Any, records: list[Any]) -> None:
    sheet.append(REVIEW_HEADERS)
    _style_header(sheet, len(REVIEW_HEADERS))
    for index, raw_record in enumerate(records, start=1):
        record = raw_record if isinstance(raw_record, dict) else {}
        excel = record.get("excel", {}) if isinstance(record.get("excel"), dict) else {}
        candidate = record.get("candidate", {}) if isinstance(record.get("candidate"), dict) else {}
        parsed = record.get("parsed_candidates", {}) if isinstance(record.get("parsed_candidates"), dict) else {}
        links = record.get("links", {}) if isinstance(record.get("links"), dict) else {}
        warnings = _string_list(record.get("warnings", []))
        row = index + 1
        sheet.append(
            [
                index,
                cell_text(record.get("code")),
                cell_text(record.get("name")),
                cell_text(excel.get("current_story")),
                cell_text(excel.get("current_backup")),
                cell_text(candidate.get("value")),
                _integer(candidate.get("length")),
                "；".join(_string_list(record.get("date_candidates", []))),
                cell_text(parsed.get("牺牲地点")),
                cell_text(parsed.get("牺牲原因")),
                "；".join(warnings),
                cell_text(candidate.get("status")),
                "",
                "",
                "",
                cell_text(links.get("mineru_text_source")),
                cell_text(links.get("correction_crop")),
            ]
        )
        _style_review_row(sheet, row, warnings, cell_text(candidate.get("status")))

    max_row = max(sheet.max_row, 2)
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(REVIEW_CONCLUSIONS) + '"',
        allow_blank=True,
    )
    validation.error = "请选择下拉列表中的审核结论。"
    validation.errorTitle = "无效审核结论"
    validation.prompt = "选择结论；需要改写时必须填写建议最终K。"
    validation.promptTitle = "人工审核"
    sheet.add_data_validation(validation)
    validation.add(f"M2:M{max_row}")

    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A1:Q{max_row}"
    sheet.sheet_view.showGridLines = False
    widths = [7, 15, 10, 34, 34, 38, 10, 24, 20, 20, 34, 18, 18, 38, 30, 38, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 32
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 90
    sheet.protection.sheet = True
    sheet.protection.autoFilter = False
    sheet.protection.sort = False
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = False


def _style_review_row(sheet: Any, row: int, warnings: list[str], status: str) -> None:
    base_fill = BLOCKED_FILL if status == "blocked" else LOCKED_FILL
    for column in range(1, len(REVIEW_HEADERS) + 1):
        cell = sheet.cell(row, column)
        cell.font = BODY_FONT
        cell.fill = base_fill
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.protection = Protection(locked=True)
    for column in (13, 14, 15):
        cell = sheet.cell(row, column)
        cell.fill = EDITABLE_FILL
        cell.border = EDITABLE_BORDER
        cell.protection = Protection(locked=False)
    if "story_candidate_long" in warnings:
        sheet.cell(row, 7).fill = WARNING_FILL
    if "multiple_date_candidates" in warnings:
        sheet.cell(row, 8).fill = WARNING_FILL
    if "story_candidate_long" in warnings or "multiple_date_candidates" in warnings:
        sheet.cell(row, 11).fill = WARNING_FILL
    for column in (1, 2, 3, 7, 12, 13):
        sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _write_source_sheet(sheet: Any, records: list[Any]) -> None:
    sheet.append(SOURCE_HEADERS)
    _style_header(sheet, len(SOURCE_HEADERS))
    for raw_record in records:
        record = raw_record if isinstance(raw_record, dict) else {}
        links = record.get("links", {}) if isinstance(record.get("links"), dict) else {}
        sheet.append(
            [
                cell_text(record.get("code")),
                cell_text(record.get("name")),
                cell_text(record.get("raw_text")),
                cell_text(links.get("mineru_text_source")),
                "；".join(_string_list(record.get("warnings", []))),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{max(sheet.max_row, 2)}"
    sheet.sheet_view.showGridLines = False
    for column, width in enumerate((15, 10, 80, 46, 46), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 110
        for cell in sheet[row]:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = LOCKED_FILL
            cell.border = CELL_BORDER
            cell.protection = Protection(locked=True)
    sheet.protection.sheet = True
    sheet.protection.autoFilter = False


def _write_instructions_sheet(sheet: Any) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "事迹候选人工审核 - 使用说明"
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 36
    instructions = [
        "1. 在“人工审核”工作表逐行审核。",
        "2. M 列选择审核结论；需要改写时，在 N 列填写建议最终 K。",
        "3. O 列记录判断依据、疑问或规则标签。",
        "4. 灰色、红色和黄色单元格为系统信息；绿色 M/N/O 列可编辑。",
        "5. 不要修改编号、姓名和系统候选；收集命令会再次校验。",
        "6. 完成后保存为 .xlsx，不要另存为 CSV。",
        "7. 本工作簿包含真实个人信息，只能保存在 data/ 等被 Git 忽略的本地目录。",
        "8. 本工具不修改 v4/v5，也不会执行 excel-apply。",
    ]
    for row, text in enumerate(instructions, start=3):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = sheet.cell(row, 1, text)
        cell.font = Font(name="Microsoft YaHei", size=11, color="202124")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="F7F8FA")
        sheet.row_dimensions[row].height = 28
    sheet.column_dimensions["A"].width = 30
    for column in "BCDE":
        sheet.column_dimensions[column].width = 18
    sheet.protection.sheet = True


def _write_metadata_sheet(sheet: Any, records: list[Any]) -> None:
    sheet.append(METADATA_HEADERS)
    for index, raw_record in enumerate(records, start=1):
        record = raw_record if isinstance(raw_record, dict) else {}
        candidate = record.get("candidate", {}) if isinstance(record.get("candidate"), dict) else {}
        sheet.append(
            [
                index + 1,
                cell_text(record.get("code")),
                cell_text(record.get("name")),
                cell_text(candidate.get("value")),
                cell_text(candidate.get("status")),
                cell_text(candidate.get("block_reason")),
                cell_text(record.get("source_stem")),
            ]
        )
    sheet.protection.sheet = True


def _style_header(sheet: Any, column_count: int) -> None:
    for column in range(1, column_count + 1):
        cell = sheet.cell(1, column)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.protection = Protection(locked=True)


def _validate_headers(sheet: Any, expected: list[str]) -> None:
    actual = [cell_text(sheet.cell(1, column).value) for column in range(1, len(expected) + 1)]
    if actual != expected:
        raise ValueError(f"Worksheet {sheet.title} headers do not match the expected review format")


def _string_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value if isinstance(item, str)]


def _integer(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0
