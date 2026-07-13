from __future__ import annotations

import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from lieshi_ocr.pipeline.audit_story import (
    build_story_candidate_audit,
    find_date_candidates,
    render_story_candidate_audit_html,
    write_story_candidate_audit_outputs,
)


def _write_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(1, 2).value = "编号"
    sheet.cell(1, 4).value = "姓名"
    sheet.cell(1, 11).value = "牺牲时间地点简要事迹"
    sheet.cell(1, 20).value = "旧K备份"
    sheet.cell(2, 2).value = "QX-0001"
    sheet.cell(2, 4).value = "张三"
    sheet.cell(2, 11).value = "当前简写"
    sheet.cell(2, 20).value = "更早历史值"
    workbook.save(path)
    workbook.close()


def _write_records(path: Path, raw_text: str) -> None:
    path.write_text(
        json.dumps(
            {
                "batch": "20260626",
                "records": [
                    {
                        "source_pdf": "data/scan/20260626/sample.pdf",
                        "source_stem": "sample",
                        "code": "QX-0001",
                        "name": "张三",
                        "fields": {
                            "牺牲时间": "1937年8月",
                            "牺牲地点": "某地",
                            "牺牲原因": "战斗中牺牲",
                            "事迹": "参加某次战斗",
                        },
                        "raw_text": raw_text,
                        "regions": {
                            "correction": {
                                "text_source": "data/work/20260626/mineru/sample.md",
                            }
                        },
                        "warnings": ["multiple_date_candidates", "needs_human_review"],
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def _write_dry_run(path: Path, candidate: str) -> None:
    path.write_text(
        json.dumps(
            {
                "proposed_changes": [],
                "blocked_changes": [
                    {
                        "id": "QX-0001:K2",
                        "code": "QX-0001",
                        "name": "张三",
                        "row": 2,
                        "column": 11,
                        "column_letter": "K",
                        "old": "当前简写",
                        "new": candidate,
                        "block_reason": "story_backup_conflict",
                        "story_length": len(candidate),
                        "warnings": ["story_backup_conflict", "story_candidate_long"],
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


class StoryCandidateAuditTests(unittest.TestCase):
    def test_finds_unique_date_candidates_without_selecting_one(self) -> None:
        self.assertEqual(
            find_date_candidates("1937 年 8 月参加工作，1942年3月15日牺牲。1937年8月再述。"),
            ["1937年8月", "1942年3月15日"],
        )

    def test_builds_read_only_blocked_story_comparison(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "v4.xlsx"
            records = root / "correction_records.json"
            dry_run = root / "dry_run_report.json"
            candidate = "候选" * 25
            raw_text = "1937年8月参加工作，1942年3月15日在某地牺牲。"
            _write_workbook(workbook)
            _write_records(records, raw_text)
            _write_dry_run(dry_run, candidate)
            workbook_before = workbook.read_bytes()

            report = build_story_candidate_audit(records, dry_run, workbook)
            item = report["records"][0]

            self.assertEqual(report["summary"]["blocked_story_count"], 1)
            self.assertEqual(report["summary"]["long_story_count"], 1)
            self.assertEqual(report["summary"]["multiple_date_count"], 1)
            self.assertEqual(item["excel"]["current_story"], "当前简写")
            self.assertEqual(item["excel"]["current_backup"], "更早历史值")
            self.assertEqual(item["candidate"]["status"], "blocked")
            self.assertEqual(item["candidate"]["value"], candidate)
            self.assertEqual(item["candidate"]["block_reason"], "story_backup_conflict")
            self.assertEqual(item["date_candidates"], ["1937年8月", "1942年3月15日"])
            self.assertEqual(item["parsed_candidates"]["牺牲地点"], "某地")
            self.assertIn("story_candidate_long", item["warnings"])
            self.assertEqual(workbook.read_bytes(), workbook_before)

    def test_html_escapes_mineru_text_and_keeps_manual_field_local(self) -> None:
        report = {
            "generated_at": "2026-07-13T12:00:00",
            "summary": {
                "record_count": 1,
                "proposed_story_count": 0,
                "blocked_story_count": 1,
                "long_story_count": 0,
                "multiple_date_count": 0,
            },
            "records": [
                {
                    "source_stem": "sample",
                    "code": "QX-0001",
                    "name": "张三",
                    "excel": {},
                    "candidate": {"status": "blocked", "value": "<img src=x onerror=alert(1)>"},
                    "date_candidates": [],
                    "parsed_candidates": {},
                    "raw_text": "<script>alert(1)</script>",
                    "links": {},
                    "warnings": [],
                }
            ],
        }

        html = render_story_candidate_audit_html(report, Path.cwd())

        self.assertNotIn("<script>alert(1)</script>", html)
        self.assertNotIn("<img src=x onerror=alert(1)>", html)
        self.assertIn("&lt;script&gt;alert(1)&lt;/script&gt;", html)
        self.assertIn("data-audit-key=\"QX-0001\"", html)
        self.assertIn("localStorage", html)

    def test_writes_json_html_and_cli_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "v4.xlsx"
            records = root / "correction_records.json"
            dry_run = root / "dry_run_report.json"
            out_dir = root / "audit"
            _write_workbook(workbook)
            _write_records(records, "1937年参加工作，1942年牺牲。")
            _write_dry_run(dry_run, "候选事迹")

            report = build_story_candidate_audit(records, dry_run, workbook)
            json_path, html_path = write_story_candidate_audit_outputs(report, out_dir)
            self.assertTrue(json_path.exists())
            self.assertTrue(html_path.exists())

            result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "lieshi_ocr.cli",
                    "audit-story",
                    "--records",
                    str(records),
                    "--dry-run",
                    str(dry_run),
                    "--base-xlsx",
                    str(workbook),
                    "--out-dir",
                    str(root / "cli-audit"),
                ],
                cwd=Path(__file__).resolve().parents[1],
                env={**os.environ, "PYTHONPATH": str(Path(__file__).resolve().parents[1] / "src")},
                text=True,
                capture_output=True,
                check=True,
            )

            self.assertIn('"blocked_story_count": 1', result.stdout)
            self.assertTrue((root / "cli-audit" / "story_candidate_audit.json").exists())
            self.assertTrue((root / "cli-audit" / "story_candidate_audit.html").exists())


if __name__ == "__main__":
    unittest.main()
