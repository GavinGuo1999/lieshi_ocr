from __future__ import annotations

import json
from pathlib import Path
import tempfile
import unittest

import openpyxl
from openpyxl.cell.rich_text import CellRichText

from lieshi_ocr.excel.apply_changes import _set_partial_red_value, apply_approved_changes
from lieshi_ocr.excel.v6 import build_v6_dry_run, write_v6_dry_run
from lieshi_ocr.excel.v6_rules import format_v6_origin, format_v6_story, format_v6_unit_role, normalize_v6_date


def _write_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ["" for _ in range(20)]
    headers[1] = "编号"
    headers[3] = "姓名"
    headers[4] = "性别"
    headers[5] = "籍贯"
    headers[6] = "出生时间"
    headers[7] = "参加革命时间"
    headers[8] = "政治面貌"
    headers[9] = "生前单位及曾任职务"
    headers[10] = "事迹"
    headers[13] = "审稿意见"
    headers[19] = "旧K备份"
    sheet.append(headers)
    row = ["" for _ in range(20)]
    row[1] = "QX-0001"
    row[3] = "张三"
    row[4] = "男"
    row[5] = "山西省晋中市祁县\n来远镇-旧村"
    row[6] = "1920"
    row[7] = "1938"
    row[8] = "群众"
    row[9] = "旧部队/战士"
    row[10] = "1940年在某地牺牲。"
    row[13] = "保留审稿意见"
    row[19] = "更早历史K"
    sheet.append(row)
    workbook.save(path)
    workbook.close()


def _item(field: str, value: str) -> dict[str, object]:
    return {
        "raw_label": field,
        "field": field,
        "value": value,
        "reason": "依据脱敏档案",
        "warnings": [],
    }


def _write_records(path: Path) -> None:
    records = [
        {
            "code": "QX-0001",
            "name": "张三",
            "source_stem": "sample",
            "raw_text": "structured",
            "warnings": [],
            "items": [
                _item("性别", "女"),
                _item("籍贯", "山西省晋中市祁县来远镇新村"),
                _item("出生时间", "1920年8月5日"),
                _item("参加革命/工作时间", "1938年7月"),
                _item("政治面貌", "中共党员"),
                _item("生前单位及曾任职务", "某部队排长"),
                _item("事迹", "1940年在某地战斗中牺牲"),
                _item("民族", "汉族"),
            ],
        },
        {
            "code": "QX-9999",
            "name": "重复人员",
            "source_stem": "duplicate",
            "raw_text": "与编号QX-0001重复，以QX-0001为准，需删除",
            "warnings": ["no_labeled_fields_found"],
            "items": [],
        },
        {
            "code": "QX-9998",
            "name": "重复人员二",
            "source_stem": "duplicate_ocr_noise",
            "raw_text": "与编号QX-0001复，以QX-0001准刑除",
            "warnings": ["no_labeled_fields_found"],
            "items": [],
        },
    ]
    path.write_text(json.dumps({"records": records}, ensure_ascii=False), encoding="utf-8")


class V6CandidateTests(unittest.TestCase):
    def test_v6_formatting_rules(self) -> None:
        self.assertEqual(normalize_v6_date("1947年8月"), "1947-8")
        self.assertEqual(normalize_v6_date("1947年8月15日"), "1947-8-15")
        self.assertEqual(normalize_v6_date("1947年"), "1947")
        self.assertEqual(normalize_v6_date("0"), "0")
        self.assertIsNone(normalize_v6_date("约1947年"))
        self.assertEqual(
            format_v6_origin("山西省晋中市祁县来远镇新村"),
            "山西省晋中市祁县\n来远镇-新村",
        )
        self.assertEqual(format_v6_unit_role("某部队排长"), "某部队/排长")
        self.assertEqual(format_v6_story("完整事迹,不截断"), "完整事迹，不截断。")

    def test_v6_dry_run_preserves_n_t_and_uses_full_story(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "v4.xlsx"
            records = root / "records.json"
            _write_workbook(workbook)
            _write_records(records)

            report = build_v6_dry_run(workbook, records)
            changes = {change["column_letter"]: change for change in report["proposed_changes"]}

            self.assertEqual(set(changes), {"E", "F", "G", "H", "I", "J", "K"})
            self.assertEqual(changes["G"]["new"], "1920-8-5")
            self.assertEqual(changes["H"]["new"], "1938-7")
            self.assertEqual(changes["K"]["new"], "1940年在某地战斗中牺牲。")
            self.assertEqual(changes["K"]["red_mode"], "partial")
            self.assertEqual(changes["G"]["red_mode"], "full")
            self.assertEqual(report["record_results"][1]["status"], "duplicate_delete_already_absent")
            self.assertEqual(report["record_results"][2]["status"], "duplicate_delete_already_absent")
            self.assertEqual(report["summary"]["unmapped_items_by_field"], {"民族": 1})

    def test_apply_writes_partial_red_runs_and_keeps_n_t(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = root / "v4.xlsx"
            records = root / "records.json"
            dry_run = root / "v6_dry_run.json"
            approved = root / "approved.json"
            candidate = root / "v6_candidate.xlsx"
            _write_workbook(base)
            _write_records(records)
            report = build_v6_dry_run(base, records)
            write_v6_dry_run(report, dry_run, root / "v6_dry_run.md")
            approved.write_text(
                json.dumps({"approved_changes": [item["id"] for item in report["proposed_changes"]]}),
                encoding="utf-8",
            )

            apply_report = apply_approved_changes(base, dry_run, approved, candidate)
            workbook = openpyxl.load_workbook(candidate, rich_text=True)
            sheet = workbook.active
            try:
                story = sheet["K2"].value
                self.assertIsInstance(story, CellRichText)
                self.assertEqual(str(story), "1940年在某地战斗中牺牲。")
                red_text = "".join(
                    block.text
                    for block in story
                    if not isinstance(block, str)
                    and block.font.color is not None
                    and block.font.color.type == "rgb"
                    and block.font.color.rgb == "FFFF0000"
                )
                self.assertIn("战斗中", red_text)
                self.assertEqual(sheet["G2"].font.color.rgb, "FFFF0000")
                self.assertEqual(sheet["N2"].value, "保留审稿意见")
                self.assertEqual(sheet["T2"].value, "更早历史K")
                self.assertEqual(sheet.row_dimensions[2].height, 50.1)
                self.assertEqual(sheet.column_dimensions["K"].width, 57.25)
            finally:
                workbook.close()
            self.assertEqual(apply_report["applied_count"], 7)

    def test_partial_red_marks_an_anchor_for_deletion_only_changes(self) -> None:
        workbook = openpyxl.Workbook()
        cell = workbook.active["A1"]
        cell.value = "保留文字多余"
        _set_partial_red_value(cell, "保留文字多余", "保留文字")

        self.assertIsInstance(cell.value, CellRichText)
        red_text = "".join(
            block.text
            for block in cell.value
            if not isinstance(block, str)
            and block.font.color is not None
            and block.font.color.type == "rgb"
            and block.font.color.rgb == "FFFF0000"
        )
        self.assertTrue(red_text)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
