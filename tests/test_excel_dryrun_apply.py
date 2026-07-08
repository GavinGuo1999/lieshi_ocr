from __future__ import annotations

import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from lieshi_ocr.excel.apply_changes import apply_approved_changes
from lieshi_ocr.excel.dry_run import build_excel_dry_run, write_dry_run_outputs
from lieshi_ocr.excel.rules import append_review_note


def _write_workbook(path: Path, review_note: str = "") -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = {
        2: "编号",
        4: "姓名",
        6: "籍贯",
        7: "出生时间",
        8: "参加革命/工作时间",
        9: "政治面貌",
        10: "生前单位及曾任职务",
        11: "牺牲时间地点简要事迹",
        14: "审稿意见",
        20: "旧K备份",
    }
    for column, value in headers.items():
        sheet.cell(1, column).value = value
    sheet.cell(2, 2).value = "QX-0001"
    sheet.cell(2, 4).value = "张三"
    sheet.cell(2, 6).value = "旧籍贯"
    sheet.cell(2, 7).value = "旧出生"
    sheet.cell(2, 8).value = "旧参加"
    sheet.cell(2, 9).value = "旧面貌"
    sheet.cell(2, 10).value = "旧单位旧职务"
    sheet.cell(2, 11).value = "旧事迹"
    sheet.cell(2, 14).value = review_note
    sheet.cell(3, 2).value = "QX-0002"
    sheet.cell(3, 4).value = "李四"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def _write_records(path: Path, name: str = "张三") -> None:
    payload = {
        "batch": "20260626",
        "text_manifest": "text_manifest.json",
        "out_dir": "review",
        "total": 2,
        "records": [
            {
                "batch": "20260626",
                "source_pdf": "sample.pdf",
                "source_stem": "sample",
                "code": "QX-0001",
                "name": name,
                "fields": {
                    "编号": "QX-0001",
                    "姓名": name,
                    "籍贯": "山西祁县",
                    "出生时间": "1920年",
                    "参加革命/工作时间": "1938年",
                    "政治面貌": "党员",
                    "生前单位及曾任职务": "某部",
                    "曾任职务": "战士",
                    "牺牲时间": "1942年",
                    "牺牲地点": "太行山",
                    "牺牲原因": "战斗牺牲",
                    "事迹": "参加抗日斗争",
                    "安葬地": "祁县",
                    "民族": "汉族",
                },
                "raw_text": "",
                "normalized_text": "",
                "regions": {},
                "warnings": ["manual_review"],
            },
            {
                "batch": "20260626",
                "source_pdf": "missing.pdf",
                "source_stem": "missing",
                "code": "QX-9999",
                "name": "王五",
                "fields": {"编号": "QX-9999", "姓名": "王五", "籍贯": "不会写入"},
                "raw_text": "",
                "normalized_text": "",
                "regions": {},
                "warnings": [],
            },
        ],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


class ExcelDryRunApplyTests(unittest.TestCase):
    def test_append_review_note_rules(self) -> None:
        self.assertEqual(append_review_note("", "new-note"), "new-note")
        self.assertEqual(append_review_note("old-note", "new-note"), "old-note\nnew-note")
        self.assertEqual(append_review_note("old-note\nnew-note", "new-note"), "old-note\nnew-note")

    def test_dry_run_builds_expected_changes_and_warnings(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "v4.xlsx"
            records = root / "correction_records.json"
            _write_workbook(xlsx)
            _write_records(records)

            report = build_excel_dry_run(xlsx, records)
            changes = {change.id: change for change in report.proposed_changes}
            result_by_code = {result.code: result for result in report.record_results}

            self.assertIn("QX-0001:F2", changes)
            self.assertIn("QX-0001:J2", changes)
            self.assertEqual(changes["QX-0001:J2"].new, "某部/战士")
            self.assertIn("QX-0001:K2", changes)
            self.assertIn("1942年在太行山战斗牺牲。", changes["QX-0001:K2"].new)
            self.assertIn("QX-0001:T2", changes)
            self.assertEqual(changes["QX-0001:T2"].new, "旧事迹")
            self.assertIn("QX-0001:N2", changes)
            self.assertIn("需人工复核", changes["QX-0001:N2"].new)
            self.assertIn("code_not_found", result_by_code["QX-9999"].warnings)

    def test_review_note_appends_to_existing_n_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "v4.xlsx"
            records = root / "correction_records.json"
            _write_workbook(xlsx, review_note="existing-note")
            _write_records(records)

            report = build_excel_dry_run(xlsx, records)
            changes = {change.id: change for change in report.proposed_changes}

            self.assertIn("QX-0001:N2", changes)
            self.assertTrue(changes["QX-0001:N2"].new.startswith("existing-note\n"))
            self.assertIn("manual_review", changes["QX-0001:N2"].new)

    def test_review_note_is_not_duplicated(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first_xlsx = root / "first.xlsx"
            records = root / "correction_records.json"
            _write_workbook(first_xlsx)
            _write_records(records)
            first_report = build_excel_dry_run(first_xlsx, records)
            first_changes = {change.id: change for change in first_report.proposed_changes}
            existing_review_note = first_changes["QX-0001:N2"].new

            second_xlsx = root / "second.xlsx"
            _write_workbook(second_xlsx, review_note=existing_review_note)
            second_report = build_excel_dry_run(second_xlsx, records)
            second_ids = [change.id for change in second_report.proposed_changes]

            self.assertNotIn("QX-0001:N2", second_ids)

    def test_name_mismatch_skips_record_changes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "v4.xlsx"
            records = root / "correction_records.json"
            _write_workbook(xlsx)
            _write_records(records, name="错误姓名")

            report = build_excel_dry_run(xlsx, records)
            ids = [change.id for change in report.proposed_changes]
            result = next(item for item in report.record_results if item.code == "QX-0001")

            self.assertNotIn("QX-0001:F2", ids)
            self.assertIn("name_mismatch", result.warnings)

    def test_apply_uses_only_approved_changes_and_marks_red(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "v4.xlsx"
            records = root / "correction_records.json"
            dry_run = root / "dry_run_report.json"
            dry_run_md = root / "dry_run_report.md"
            approved = root / "approved_changes.json"
            output = root / "candidate.xlsx"
            _write_workbook(xlsx)
            _write_records(records)
            report = build_excel_dry_run(xlsx, records)
            write_dry_run_outputs(report, dry_run, dry_run_md)
            approved.write_text(
                json.dumps({"approved_changes": ["QX-0001:F2", {"id": "QX-0001:T2", "approved": True}]}),
                encoding="utf-8",
            )

            apply_report = apply_approved_changes(xlsx, dry_run, approved, output)
            workbook = openpyxl.load_workbook(output)
            sheet = workbook.active
            try:
                self.assertEqual(sheet.cell(2, 6).value, "山西祁县")
                self.assertEqual(sheet.cell(2, 20).value, "旧事迹")
                self.assertEqual(sheet.cell(2, 10).value, "旧单位旧职务")
                self.assertEqual(sheet.cell(2, 6).font.color.rgb, "FFFF0000")
                self.assertEqual(apply_report["applied_count"], 2)
            finally:
                workbook.close()

            original = openpyxl.load_workbook(xlsx)
            try:
                self.assertEqual(original.active.cell(2, 6).value, "旧籍贯")
            finally:
                original.close()

    def test_cli_excel_dry_run_and_apply(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "v4.xlsx"
            records = root / "correction_records.json"
            out_dir = root / "excel"
            approved = root / "approved_changes.json"
            output = root / "candidate.xlsx"
            _write_workbook(xlsx)
            _write_records(records)

            env = {**os.environ, "PYTHONPATH": str(Path(__file__).resolve().parents[1] / "src")}
            subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "lieshi_ocr.cli",
                    "excel-dry-run",
                    "--base-xlsx",
                    str(xlsx),
                    "--records",
                    str(records),
                    "--out-dir",
                    str(out_dir),
                ],
                cwd=Path(__file__).resolve().parents[1],
                env=env,
                text=True,
                capture_output=True,
                check=True,
            )
            self.assertTrue((out_dir / "dry_run_report.json").exists())
            self.assertTrue((out_dir / "dry_run_report.md").exists())

            approved.write_text(json.dumps({"approved_changes": ["QX-0001:F2"]}), encoding="utf-8")
            subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "lieshi_ocr.cli",
                    "excel-apply",
                    "--base-xlsx",
                    str(xlsx),
                    "--dry-run",
                    str(out_dir / "dry_run_report.json"),
                    "--approved",
                    str(approved),
                    "--out-xlsx",
                    str(output),
                ],
                cwd=Path(__file__).resolve().parents[1],
                env=env,
                text=True,
                capture_output=True,
                check=True,
            )
            self.assertTrue(output.exists())
            self.assertTrue(output.with_suffix(".apply_report.json").exists())


if __name__ == "__main__":
    unittest.main()
