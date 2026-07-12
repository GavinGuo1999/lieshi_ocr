from __future__ import annotations

import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from lieshi_ocr.pipeline.audit_ocr import (
    build_ocr_audit,
    clean_ocr_name,
    render_ocr_audit_html,
    write_ocr_audit_outputs,
)


def _write_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(1, 2).value = "编号"
    sheet.cell(1, 4).value = "姓名"
    sheet.cell(2, 2).value = "QX-0001"
    sheet.cell(2, 4).value = "张三"
    workbook.save(path)
    workbook.close()


def _write_manifests(root: Path, name_text: str = "姓名：张 三。", record_name: str = "张三") -> tuple[Path, Path, Path]:
    code_crop = root / "crop" / "sample__code.pdf"
    name_crop = root / "crop" / "sample__name.pdf"
    correction_crop = root / "crop" / "sample__correction.pdf"
    mineru_text = root / "mineru" / "sample.md"
    for path in (code_crop, name_crop, correction_crop):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"%PDF-1.4\n")
    mineru_text.parent.mkdir(parents=True, exist_ok=True)
    mineru_text.write_text("1937年在某地牺牲。", encoding="utf-8")

    text_manifest = root / "text_manifest.json"
    text_manifest.write_text(
        json.dumps(
            {
                "batch": "20260626",
                "records": [
                    {
                        "source_pdf": (root / "sample.pdf").as_posix(),
                        "source_stem": "sample",
                        "region": "code",
                        "crop_pdf": code_crop.as_posix(),
                        "engine": "rapidocr",
                        "text": "QX-0001",
                        "confidence": 0.97,
                        "warnings": [],
                        "text_source": "",
                    },
                    {
                        "source_pdf": (root / "sample.pdf").as_posix(),
                        "source_stem": "sample",
                        "region": "name",
                        "crop_pdf": name_crop.as_posix(),
                        "engine": "rapidocr",
                        "text": name_text,
                        "confidence": 0.88,
                        "warnings": [],
                        "text_source": "",
                    },
                    {
                        "source_pdf": (root / "sample.pdf").as_posix(),
                        "source_stem": "sample",
                        "region": "correction",
                        "crop_pdf": correction_crop.as_posix(),
                        "engine": "mineru_text",
                        "text": "1937年在某地牺牲。",
                        "confidence": 1.0,
                        "warnings": [],
                        "text_source": mineru_text.as_posix(),
                    },
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    records = root / "correction_records.json"
    records.write_text(
        json.dumps(
            {
                "batch": "20260626",
                "records": [
                    {
                        "source_pdf": (root / "sample.pdf").as_posix(),
                        "source_stem": "sample",
                        "code": "QX-0001",
                        "name": record_name,
                        "fields": {"编号": "QX-0001", "姓名": record_name, "事迹": "某地牺牲"},
                        "raw_text": "1937年在某地牺牲。",
                        "regions": {},
                        "warnings": ["needs_human_review"],
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return text_manifest, records, mineru_text


class OcrAuditReportTests(unittest.TestCase):
    def test_safe_name_cleanup_only_removes_labels_whitespace_and_punctuation(self) -> None:
        self.assertEqual(clean_ocr_name(" 姓名： 张 三。\n"), "张三")
        self.assertEqual(clean_ocr_name("章三"), "章三")

    def test_builds_match_diagnostics_and_readable_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "v4.xlsx"
            _write_workbook(workbook)
            text_manifest, records, mineru_text = _write_manifests(root)
            workbook_before = workbook.read_bytes()

            report = build_ocr_audit(text_manifest, records, workbook)
            json_path, html_path = write_ocr_audit_outputs(report, root / "audit")
            item = report["records"][0]
            html = html_path.read_text(encoding="utf-8")

            self.assertEqual(report["summary"]["name_status_counts"], {"match": 1})
            self.assertTrue(item["code"]["excel_exists"])
            self.assertEqual(item["code"]["excel_row"], 2)
            self.assertEqual(item["name"]["cleaned"], "张三")
            self.assertEqual(item["name"]["excel_expected"], "张三")
            self.assertEqual(item["links"]["mineru_text_source"], mineru_text.as_posix())
            self.assertEqual(workbook.read_bytes(), workbook_before)
            self.assertTrue(json_path.exists())
            self.assertIn("OCR 审计报告", html)
            self.assertIn("MinerU 原文", html)
            self.assertIn("人工结论", html)

    def test_reports_exact_name_mismatch_without_fuzzy_release(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "v4.xlsx"
            _write_workbook(workbook)
            text_manifest, records, _ = _write_manifests(root, name_text="章三", record_name="章三")

            report = build_ocr_audit(text_manifest, records, workbook)
            item = report["records"][0]

            self.assertEqual(item["name"]["match_status"], "name_mismatch")
            self.assertEqual(item["name"]["excel_expected"], "张三")
            self.assertFalse(item["name"]["cleaned_matches_excel"])
            self.assertIn("name_mismatch", item["warnings"])

    def test_html_escapes_untrusted_ocr_text(self) -> None:
        report = {
            "generated_at": "2026-07-12T12:00:00",
            "summary": {"record_count": 1, "code_found": 0, "name_status_counts": {}},
            "records": [
                {
                    "source_stem": "sample",
                    "code": {},
                    "name": {},
                    "links": {},
                    "fields": {"事迹": "<img src=x onerror=alert(1)>"},
                    "raw_text": "<script>alert(1)</script>",
                    "warnings": [],
                }
            ],
        }

        html = render_ocr_audit_html(report, Path.cwd())

        self.assertNotIn("<script>alert(1)</script>", html)
        self.assertNotIn("<img src=x onerror=alert(1)>", html)
        self.assertIn("&lt;script&gt;alert(1)&lt;/script&gt;", html)

    def test_cli_writes_json_and_html_to_explicit_temp_directory(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "v4.xlsx"
            out_dir = root / "audit"
            _write_workbook(workbook)
            text_manifest, records, _ = _write_manifests(root)

            result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "lieshi_ocr.cli",
                    "audit-ocr",
                    "--text-manifest",
                    str(text_manifest),
                    "--records",
                    str(records),
                    "--base-xlsx",
                    str(workbook),
                    "--out-dir",
                    str(out_dir),
                ],
                cwd=Path(__file__).resolve().parents[1],
                env={**os.environ, "PYTHONPATH": str(Path(__file__).resolve().parents[1] / "src")},
                text=True,
                capture_output=True,
                check=True,
            )

            self.assertIn('"record_count": 1', result.stdout)
            self.assertTrue((out_dir / "ocr_audit_report.json").exists())
            self.assertTrue((out_dir / "ocr_audit_report.html").exists())


if __name__ == "__main__":
    unittest.main()
