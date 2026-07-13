from __future__ import annotations

import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from lieshi_ocr.pipeline.story_review_xlsx import (
    METADATA_SHEET,
    REVIEW_SHEET,
    SOURCE_SHEET,
    collect_story_review_decisions,
    write_story_review_decisions,
    write_story_review_workbook,
)


def _audit_report() -> dict:
    return {
        "generated_at": "2026-07-13T12:00:00",
        "records": [
            {
                "source_stem": "sample-1",
                "code": "QX-0001",
                "name": "张三",
                "excel": {"current_story": "当前简写一", "current_backup": "历史备份一"},
                "candidate": {"status": "blocked", "value": "候选事迹一" * 10, "length": 50},
                "date_candidates": ["1937年8月", "1942年3月15日"],
                "parsed_candidates": {"牺牲地点": "某地", "牺牲原因": "战斗中牺牲"},
                "raw_text": "完整 MinerU 正文一，不得截断。",
                "links": {
                    "mineru_text_source": "data/work/sample-1.md",
                    "correction_crop": "data/work/sample-1__correction.pdf",
                },
                "warnings": [
                    "story_backup_conflict",
                    "story_candidate_long",
                    "multiple_date_candidates",
                    "needs_human_review",
                ],
            },
            {
                "source_stem": "sample-2",
                "code": "QX-0002",
                "name": "李四",
                "excel": {"current_story": "当前简写二", "current_backup": ""},
                "candidate": {"status": "proposed", "value": "候选事迹二", "length": 5},
                "date_candidates": ["1943年"],
                "parsed_candidates": {"牺牲地点": "另一地", "牺牲原因": "因伤牺牲"},
                "raw_text": "完整 MinerU 正文二。",
                "links": {
                    "mineru_text_source": "data/work/sample-2.md",
                    "correction_crop": "data/work/sample-2__correction.pdf",
                },
                "warnings": ["needs_human_review"],
            },
        ],
    }


class StoryReviewXlsxTests(unittest.TestCase):
    def test_generates_review_workbook_layout_validation_and_protection(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "review.xlsx"
            summary = write_story_review_workbook(_audit_report(), output)
            workbook = openpyxl.load_workbook(output)
            try:
                self.assertEqual(summary["record_count"], 2)
                self.assertEqual(workbook.sheetnames[:3], ["人工审核", "原始正文", "使用说明"])
                self.assertEqual(workbook[METADATA_SHEET].sheet_state, "veryHidden")
                review = workbook[REVIEW_SHEET]
                self.assertEqual(review.freeze_panes, "D2")
                self.assertEqual(review.auto_filter.ref, "A1:Q3")
                self.assertTrue(review.protection.sheet)
                self.assertTrue(review["B2"].protection.locked)
                self.assertFalse(review["M2"].protection.locked)
                self.assertFalse(review["N2"].protection.locked)
                self.assertFalse(review["O2"].protection.locked)
                validations = list(review.data_validations.dataValidation)
                self.assertEqual(len(validations), 1)
                self.assertIn("M2:M3", str(validations[0].sqref))
                self.assertIn("需要改写", validations[0].formula1)
                self.assertEqual(review["F2"].value, "候选事迹一" * 10)
                self.assertEqual(review["Q2"].value, "data/work/sample-1__correction.pdf")
                source = workbook[SOURCE_SHEET]
                self.assertEqual(source["C2"].value, "完整 MinerU 正文一，不得截断。")
                self.assertEqual(source.freeze_panes, "A2")
                self.assertEqual(review["G2"].fill.fgColor.rgb[-6:], "FFF2CC")
                self.assertEqual(review["M2"].fill.fgColor.rgb[-6:], "E2F0D9")
            finally:
                workbook.close()

    def test_collects_pass_rewrite_and_blocking_decisions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            review_path = root / "review.xlsx"
            output_json = root / "decisions.json"
            write_story_review_workbook(_audit_report(), review_path)
            workbook = openpyxl.load_workbook(review_path)
            review = workbook[REVIEW_SHEET]
            review["M2"] = "通过系统候选"
            review["O2"] = "候选可接受"
            review["M3"] = "需要改写"
            review["N3"] = "人工改写后的事迹"
            review["O3"] = "删除冗余日期"
            workbook.save(review_path)
            workbook.close()

            report = collect_story_review_decisions(review_path)
            write_story_review_decisions(report, output_json)

            self.assertEqual(report["summary"]["reviewed_count"], 2)
            self.assertEqual(report["summary"]["passed_count"], 1)
            self.assertEqual(report["summary"]["rewritten_count"], 1)
            self.assertEqual(report["summary"]["blocked_count"], 0)
            self.assertEqual(report["decisions"][0]["final_story"], "候选事迹一" * 10)
            self.assertEqual(report["decisions"][1]["final_story"], "人工改写后的事迹")
            self.assertTrue(report["decisions"][0]["content_approved"])
            self.assertTrue(report["decisions"][0]["requires_backup_resolution"])
            self.assertFalse(report["decisions"][0]["approvable"])
            self.assertTrue(report["decisions"][1]["approvable"])
            self.assertTrue(output_json.exists())

            workbook = openpyxl.load_workbook(review_path)
            workbook[REVIEW_SHEET]["M2"] = "信息不足"
            workbook[REVIEW_SHEET]["M3"] = None
            workbook.save(review_path)
            workbook.close()
            blocked_report = collect_story_review_decisions(review_path)
            self.assertEqual(blocked_report["summary"]["blocked_count"], 1)
            self.assertFalse(blocked_report["decisions"][0]["approvable"])
            self.assertFalse(blocked_report["decisions"][0]["content_approved"])
            self.assertEqual(blocked_report["decisions"][0]["final_story"], "")

    def test_rewrite_requires_final_story(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            review_path = Path(tmp) / "review.xlsx"
            write_story_review_workbook(_audit_report(), review_path)
            workbook = openpyxl.load_workbook(review_path)
            workbook[REVIEW_SHEET]["M2"] = "需要改写"
            workbook.save(review_path)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "rewritten final story is required"):
                collect_story_review_decisions(review_path)

    def test_collect_rejects_changed_identity_or_system_candidate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            review_path = Path(tmp) / "review.xlsx"
            write_story_review_workbook(_audit_report(), review_path)
            workbook = openpyxl.load_workbook(review_path)
            workbook[REVIEW_SHEET]["B2"] = "CHANGED"
            workbook.save(review_path)
            workbook.close()
            with self.assertRaisesRegex(ValueError, "code or name differs"):
                collect_story_review_decisions(review_path)

            write_story_review_workbook(_audit_report(), review_path)
            workbook = openpyxl.load_workbook(review_path)
            workbook[REVIEW_SHEET]["F2"] = "CHANGED"
            workbook.save(review_path)
            workbook.close()
            with self.assertRaisesRegex(ValueError, "system story candidate differs"):
                collect_story_review_decisions(review_path)

    def test_collect_cli_writes_summary_without_touching_review_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            review_path = root / "review.xlsx"
            output_json = root / "decisions.json"
            write_story_review_workbook(_audit_report(), review_path)
            workbook = openpyxl.load_workbook(review_path)
            workbook[REVIEW_SHEET]["M2"] = "暂不处理"
            workbook.save(review_path)
            workbook.close()
            before = review_path.read_bytes()

            result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "lieshi_ocr.cli",
                    "collect-story-review",
                    "--review-xlsx",
                    str(review_path),
                    "--out-json",
                    str(output_json),
                ],
                cwd=Path(__file__).resolve().parents[1],
                env={**os.environ, "PYTHONPATH": str(Path(__file__).resolve().parents[1] / "src")},
                text=True,
                capture_output=True,
                check=True,
            )

            self.assertIn('"reviewed_count": 1', result.stdout)
            self.assertTrue(output_json.exists())
            self.assertEqual(review_path.read_bytes(), before)


if __name__ == "__main__":
    unittest.main()
